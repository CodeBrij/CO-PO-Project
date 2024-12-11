import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import Workbook

# Function to create a thin border
def create_border():
    thin = Side(border_style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# User inputs
subject = input("Enter the subject name: ")
total_roll = int(input("Enter the total number of roll numbers: "))
LabTarget = float(input("Enter the Lab Target: "))
lab_type = input("Enter the lab type ('non-group' or 'group'): ").strip().lower()

# Create workbook and sheets
workbook = Workbook()
orals_sheet = workbook.active
orals_sheet.title = "Orals"

# Common setup for Orals
orals_sheet['A1'] = f"{subject} Orals"
orals_sheet['A1'].font = Font(size=14, bold=True)  # Make the heading bold and larger
orals_sheet['A1'].alignment = Alignment(horizontal='center')  # Center align the heading
orals_sheet['A1'].border = create_border()  # Add border to heading
orals_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
orals_sheet['A2'] = f"Target = {LabTarget}%"
orals_sheet['A2'].border = create_border()  # Add border to Lab Target
orals_sheet['A3'] = "Roll No."
orals_sheet['A3'].border = create_border()  # Add border to header
orals_sheet['B3'] = "Name"
orals_sheet['B3'].border = create_border()  # Add border to header
orals_sheet['C3'] = "Marks(25)"
orals_sheet['C3'].border = create_border()  # Add border to header

for i in range(total_roll):
    cell = orals_sheet[f'A{i+4}']
    cell.value = i + 1
    cell.border = create_border()  # Add border to roll number cells

endCol = i + 4

footer_info = [
    ("Count(appeared)", f'B{endCol+2}'),
    (f"Count(>={LabTarget}%)", f'B{endCol+3}'),
    (f"% count(>={LabTarget}%) w.r.t appeared", f'B{endCol+4}'),
    ("AL (All Los)", f'B{endCol+5}')
]

for text, position in footer_info:
    cell = orals_sheet[position]
    cell.value = text
    cell.border = create_border()  # Add border to footer cells

# Non-Group Wise Template
if lab_type == 'non-group':
    total_exp = int(input("Enter the total number of experiments: "))
    LOs = [input(f'Enter the LO for Exp{i+1}: ') for i in range(total_exp)]

    lab_sheet = workbook.create_sheet(title="Lab")
    lab_sheet['A1'] = f"{subject} Lab Work - Ungrouped"
    lab_sheet['A1'].font = Font(size=14, bold=True)
    lab_sheet['A1'].alignment = Alignment(horizontal='center')
    lab_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_exp+6)
    lab_sheet['A2'] = f"Target = {LabTarget}"

    lab_sheet['A3'] = "Roll No."
    lab_sheet['B3'] = "Name"
    lab_sheet['B2'] = f"Total Experiment = {total_exp}"
    lab_sheet.insert_rows(4)
    lab_sheet['A5'] = "LO"
    lab_sheet.merge_cells('A5:B5')
    lab_sheet['A5'].font = Font(size=12, bold=True)
    lab_sheet['A5'].alignment = Alignment(horizontal='center')

    for i in range(total_exp):
        lab_sheet.cell(row=4, column=i+3, value=f"Exp {i+1}")
        lab_sheet.cell(row=5, column=i+3, value=LOs[i])

    lab_sheet.cell(row=4, column=total_exp+3, value="Average(15)")
    lab_sheet.cell(row=4, column=total_exp+4, value="Assignment 1")
    lab_sheet.cell(row=4, column=total_exp+5, value="Assignment 2")
    lab_sheet.cell(row=4, column=total_exp+6, value="Average Assignment(5)")

    for i in range(6, total_roll + 6):
        lab_sheet[f'A{i}'] = i - 5

    lab_sheet.column_dimensions['A'].width = 10
    lab_sheet.column_dimensions['B'].width = 25

    roll_end = total_roll + 5
    lab_sheet[f'A{roll_end+2}'] = f'Count>={LabTarget}%'
    lab_sheet[f'A{roll_end+3}'] = f'%Count'
    lab_sheet[f'A{roll_end+4}'] = f'AL'

    for row in lab_sheet.iter_rows(min_row=3, max_row=roll_end+6, min_col=1, max_col=total_exp+6):
        for cell in row:
            cell.border = create_border()
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Group Wise Template
elif lab_type == 'group':
    groupSize = int(input("Enter the size of group: "))
    criteria = int(input("Enter the number of criteria for marks: "))
    critList = [input(f"Enter criteria {i + 1}: ") for i in range(criteria)]
    loList = [input(f"Enter the LO for criteria {i + 1}: ") for i in range(criteria)]

    lab_sheet = workbook.create_sheet(title="Lab")
    lab_sheet.merge_cells('A1:E1')
    lab_sheet['A1'] = f"{subject} Lab Work - Grouped"
    lab_sheet['A1'].font = Font(size=14, bold=True)
    lab_sheet['A1'].alignment = Alignment(horizontal='center')
    lab_sheet['A2'] = f"Target={LabTarget}"
    lab_sheet['A3'] = "Group No."
    lab_sheet['B3'] = "Roll No."
    lab_sheet['C3'] = "Name of Student"
    lab_sheet['D3'] = "Project Name"
    lab_sheet['A4'] = "LOs Mapped"
    lab_sheet.column_dimensions['C'].width = 30

    startCell = 5
    groupCount = 1

    for roll_no in range(1, total_roll + 1):
        current_row = startCell + roll_no - 1
        if (roll_no - 1) % groupSize == 0:
            lab_sheet[f"A{current_row}"] = groupCount
            lab_sheet.merge_cells(f'A{current_row}:A{min(current_row + groupSize - 1, startCell + total_roll - 1)}')
            lab_sheet.merge_cells(f'D{current_row}:D{min(current_row + groupSize - 1, startCell + total_roll - 1)}')
            groupCount += 1
        lab_sheet[f"B{current_row}"] = roll_no
        lab_sheet[f"C{current_row}"] = f"Student {roll_no}"

    for i in range(criteria):
        lab_sheet.cell(row=3, column=5 + i, value=critList[i])
        lab_sheet.cell(row=4, column=5 + i, value=loList[i])

    current_row = startCell + total_roll - 1
    lab_sheet[f'A{current_row+2}'] = f"Count>={LabTarget}%"
    lab_sheet[f'A{current_row+3}'] = f"%Count"
    lab_sheet[f'A{current_row+4}'] = "AL"

    for row in lab_sheet.iter_rows(min_row=3, max_row=current_row + 4, min_col=1, max_col=5 + criteria):
        for cell in row:
            cell.border = create_border()

print(f'C{current_row}')
print(f'A{current_row+2}')
print(f'A{current_row+3}')
print(f'A{current_row+4}')

# Save the workbook
workbook.save(f"{subject}_Lab_Records.xlsx")
