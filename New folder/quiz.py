import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook("quiz.xlsx")

sheet = wb['Sheet1']

char_quiz = 'A'
quizRow = 1

target = 52.5

# Find the first empty row in column A
for row in range(1, 100):
    if sheet[f'{char_quiz}{row}'].value is None:
        quizRow = row
        break

# Merge cells in columns A and B for four rows starting at quizRow
for i in range(0, 4):
    sheet.merge_cells(f'A{quizRow + i}:B{quizRow + i}')

# Set the values in the merged cells
sheet[f'A{quizRow}'] = 'Count(Attempted)'
sheet[f'A{quizRow + 1}'] = f'Count(>={target}%)'
sheet[f'A{quizRow + 2}'] = f'% Count(>={target}%)'
sheet[f'A{quizRow + 3}'] = f'AL(Based on >={target}% count)'

rollNoRow = 0
COrow = 0
# Find the rows for Roll No and CO
for row in range(1, 4):
    if sheet[f'{char_quiz}{row}'].value and 'CO' in str(sheet[f'{char_quiz}{row}'].value):
        COrow = row + 1
        rollNoRow = row + 1
        break

# Merge cells for "Attainment Level" and "CO"
merge1 = f'C{quizRow + 6}'
merge2 = f'D{quizRow + 6}'
sheet.merge_cells(f'{merge1}:{merge2}')
sheet[f'C{quizRow + 6}'] = 'Attainment Level'
sheet[f'C{quizRow + 7}'] = 'CO'
sheet[f'C{quizRow + 7}'].font = Font(b=True)

# Set CO labels
for i in range(0, 6):
    sheet[f'C{quizRow + 8 + i}'] = f'CO{i + 1}'

columnLetter = []
# Find columns with non-empty headers
for col in range(3, 60):
    colLetter = openpyxl.utils.get_column_letter(col)
    if sheet[f'{colLetter}1'].value is not None:
        columnLetter.append(colLetter)
    else:
        break

# Set formulas in the appropriate cells
for index in range(0, len(columnLetter)):
    sheet[f'{columnLetter[index]}{quizRow}'] = f'=COUNTA({columnLetter[index]}{rollNoRow}:{columnLetter[index]}{quizRow - 1})'
    sheet[f'{columnLetter[index]}{quizRow + 1}'] = f'=COUNTIF({columnLetter[index]}{rollNoRow}:{columnLetter[index]}{quizRow - 1}, ">={(target/100)*1}")'
    sheet[f'{columnLetter[index]}{quizRow + 2}'] = f'=ROUND(({columnLetter[index]}{quizRow + 1}/{columnLetter[index]}{quizRow}*100), 1)'
    sheet[f'{columnLetter[index]}{quizRow + 3}'] = f'=IF({columnLetter[index]}{quizRow + 2}<60,1,IF(AND({columnLetter[index]}{quizRow + 2}>59,{columnLetter[index]}{quizRow + 2}<70),2,IF(AND({columnLetter[index]}{quizRow + 2}>69,{columnLetter[index]}{quizRow + 2}<80),3,4)))'

CO1, CO2, CO3, CO4, CO5, CO6 = [], [], [], [], [], []

# Extract CO column references
for col in columnLetter:
    cell = sheet[f'{col}{COrow}']
    if cell.value:
        cell_value = str(cell.value)
        if 'CO' in cell_value:
            co_index = cell_value.index('CO')
            co_numbers = ''
            for char in cell_value[co_index + 2:]:
                if char.isdigit() or char == ',':
                    co_numbers += char
                else:
                    break
            values = [int(val.strip()) for val in co_numbers.split(',') if val.strip().isdigit()]
        else:
            values = [int(val.strip()) for val in cell_value.split(',') if val.strip().isdigit()]

        if 1 in values:
            CO1.append(col)
        if 2 in values:
            CO2.append(col)
        if 3 in values:
            CO3.append(col)
        if 4 in values:
            CO4.append(col)
        if 5 in values:
            CO5.append(col)
        if 6 in values:
            CO6.append(col)

# Calculate average formulas for each CO
if CO1:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO1])}),1)"
    sheet[f'D{quizRow + 8}'] = average_formula
else:
    sheet[f'D{quizRow + 8}'] = '-'

if CO2:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO2])}),1)"
    sheet[f'D{quizRow + 9}'] = average_formula
else:
    sheet[f'D{quizRow + 9}'] = '-'

if CO3:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO3])}),1)"
    sheet[f'D{quizRow + 10}'] = average_formula
else:
    sheet[f'D{quizRow + 10}'] = '-'

if CO4:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO4])}),1)"
    sheet[f'D{quizRow + 11}'] = average_formula
else:
    sheet[f'D{quizRow + 11}'] = '-'

if CO5:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO5])}),1)"
    sheet[f'D{quizRow + 12}'] = average_formula
else:
    sheet[f'D{quizRow + 12}'] = '-'

if CO6:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO6])}),1)"
    sheet[f'D{quizRow + 13}'] = average_formula
else:
    sheet[f'D{quizRow + 13}'] = '-'

wb.save('quiz.xlsx')
