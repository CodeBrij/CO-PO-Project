    
import re
from openpyxl.styles import *
from openpyxl import *
from openpyxl.styles import Border, Side

# Load the workbook and select the sheet
workbook = load_workbook('./DSA_Lab.xlsx')
sheet = workbook['Lab']

# Find the roll count by identifying the row where "AL" is located
rollCount = 0
for i in range(1, 200):  # Corrected to iterate over a range of rows
    if sheet[f'A{i}'].value == "AL":
        rollCount = i
        break  # Stop once we find "AL"

print("RolCount", rollCount)

#alphabet array
alphabets = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

total_roll = rollCount - 8 #total_students
startRow = 5 #start of marks row
endRow = rollCount - 4 #end of marks row
calRow = rollCount - 2 #calculation row started

startCol = 'E'
endCol = 0

for col in range(1, sheet.max_column + 1):
    if sheet.cell(row=3, column=col).value == None:
        endCol = col - 1

# Get the target value from cell A2
target = sheet['A2'].value
match = re.search(r'\d+(\.\d+)?', target)  # Updated to match integers and floats
if match:
    targetvalue = float(match.group())

print("target:",targetvalue)



# Find the last column for processing (where data ends)
endCol = None
for j in range(1, 26):  # Iterate through columns in the sheet
    if sheet.cell(row=3, column=j).value is None:  # Find the first empty column in row 3
        endCol = j - 2  # The last filled column
        break

# Getting the LOs from the LOs row

los_dict = {}
    
for col in range(ord(startCol), ord(alphabets[endCol]) + 1):
        current_col_letter = chr(col)
        cell_value = sheet[f'{current_col_letter}4'].value
        
        if cell_value:  # If the cell has a value
            values = cell_value.split(',')
            for value in values:
                value = value.strip()  # Remove any leading/trailing whitespace
                if value.isnumeric():  # Ensure it's a numeric value
                    lo_key = f'LO{value}'
                    if lo_key not in los_dict:
                        los_dict[lo_key] = []
                    los_dict[lo_key].append(current_col_letter)
    
print(los_dict)

# Calc. values at End Col
for col in range(ord(startCol), ord(alphabets[endCol]) + 1):
    current_col_letter = chr(col)
    sheet[f'{current_col_letter}{calRow}'] = f'=COUNTIF({current_col_letter}{startRow}:{current_col_letter}{endRow},">={5*targetvalue/100}")' 
    sheet[f'{current_col_letter}{calRow+1}'] = f'=ROUND((({current_col_letter}{calRow}/{total_roll})*100),1)' 
    sheet[f'{current_col_letter}{calRow+2}'] = f'=IF({current_col_letter}{calRow+1}<60,1,IF(AND({current_col_letter}{calRow+1}>59,{current_col_letter}{calRow+1}<70),2,IF(AND({current_col_letter}{calRow+1}>69,{current_col_letter}{calRow+1}<80),3,4)))'

new_row = calRow + 4

# Define the border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define the font style for bold text
bold_font = Font(bold=True)

# Set the headers
sheet[f'B{new_row}'] = "LOs"
sheet[f'C{new_row}'] = "AL"

# Apply the border and bold font to the headers
sheet[f'B{new_row}'].border = thin_border
sheet[f'B{new_row}'].font = bold_font

sheet[f'C{new_row}'].border = thin_border
sheet[f'C{new_row}'].font = bold_font

i = 1

# Iterate over each key in los_dict to create and write the formulas
for key in los_dict.keys():
    columns = los_dict[key]  # Get the list of columns for the current LO
    column_ranges = ','.join([f'{col}{calRow+2}' for col in columns])  # Create the range for AVERAGE formula
    sheet[f'B{new_row+1}'] = key
    sheet[f'C{new_row+1}'] = f'=ROUND(AVERAGE({column_ranges}),1)'  # AVERAGE formula

    # Add border to the cells
    sheet[f'B{new_row+1}'].border = thin_border
    sheet[f'C{new_row+1}'].border = thin_border
    
    new_row += 1  # Increment new_row for the next set of entries


# Save the workbook after adding the new table
workbook.save('IoE_Lab_Updated_with_LO_AL_Table.xlsx')