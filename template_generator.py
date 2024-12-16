import openpyxl
from openpyxl.styles import Alignment,Font
from openpyxl.styles.borders import Border,Side
from openpyxl import Workbook
from tkinter import filedialog
from tkinter import messagebox
import tkinter as tk
import os
from pathlib import Path
import customtkinter as ctk
from CTkMessagebox import CTkMessagebox
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

def template_gen(coTextArray,basic_values_temp,midSem_Co_values_temp,CA1_Co_arr_temp,CA2_Co_arr_temp,CA3_Co_arr_temp,al_values_temp, receiversEmail):
    workbook=Workbook() 
    print("Number of CAs:", basic_values_temp[10])
    cosCount = (int)(basic_values_temp[14])  #no. of COs for the subject
    print(f"Cos Count = no. of COs for the subject = {cosCount}")
    sheet1 = workbook.active
    sheet1.title = "Midsem"
    sheet2=workbook.create_sheet(title="Endsem")
    sheet3=workbook.create_sheet(title="CA1")
    sheet4=workbook.create_sheet(title="CA2")
    if basic_values_temp[10]=="3":
        sheet7=workbook.create_sheet(title="CA3")
        print("Created")
    sheet5=workbook.create_sheet(title="Survey")
    sheet6=workbook.create_sheet(title='Attainment')
    sheet8=workbook.create_sheet(title='PO Attainment')
    sheet0=workbook.create_sheet(title="CO Information")
    
    sheet1.column_dimensions['A'].width = 42
    sheet1.merge_cells("A1:O1")
    sheet1["A1"].value="Vivekanand Education Society's Institute of Technology"

    
    sheet1.merge_cells("A2:O2")
    sheet1["A2"].value="Department of "+basic_values_temp[1]+""

    sheet1.merge_cells("A3:O3")
    sheet1["A3"].value="Academic Year :"+basic_values_temp[5]+""
    
    sheet1.merge_cells("A4:O4")
    
    sheet1.merge_cells("A5:O5")
    sheet1["A5"].value="  Subject : "+basic_values_temp[4]+"                                                                                                                                                                       Class : "+basic_values_temp[7]+""


    sheet1.merge_cells("A6:O6")
    sheet1["A6"].value="  Subject Teacher :"+basic_values_temp[6]+"                                                                                                                                                                Semester : "+basic_values_temp[3]+""
  
    

    sheet1.merge_cells("A7:O7")
    sheet1["A7"].value="Number of Students ="+basic_values_temp[0]+""

    sheet1["A8"]="Roll Nos."
    sheet1["B8"]="1a"
    sheet1["C8"]="1b"
    sheet1["D8"]="1c"
    sheet1["E8"]="1d"
    sheet1["F8"]="1e"
    sheet1["G8"]="1f"
    sheet1["H8"]="Q1"
    sheet1["I8"]="2a"
    sheet1["J8"]="2b"
    sheet1["K8"]="Q2"
    sheet1["L8"]="3a"
    sheet1["M8"]="3b"
    sheet1["N8"]="Q3"
    sheet1["O8"]="Total"

    sheet1["A9"]="COs"
    sheet1["B9"]="CO"+midSem_Co_values_temp[0]+""
    sheet1["C9"]="CO"+midSem_Co_values_temp[1]+""
    sheet1["D9"]="CO"+midSem_Co_values_temp[2]+""
    sheet1["E9"]="CO"+midSem_Co_values_temp[3]+""
    sheet1["F9"]="CO"+midSem_Co_values_temp[4]+""
    sheet1["G9"]="CO"+midSem_Co_values_temp[5]+""
    # sheet1["H9"]=""
    sheet1["I9"]="CO"+midSem_Co_values_temp[6]+""
    sheet1["J9"]="CO"+midSem_Co_values_temp[7]+""
    # sheet1["K9"]=""
    sheet1["L9"]="CO"+midSem_Co_values_temp[8]+""
    sheet1["M9"]="CO"+midSem_Co_values_temp[9]+""
    # sheet1["N9"]=""
    sheet1["O9"]=20

    total_roll=int(basic_values_temp[0])
    for i in range(1,total_roll+1):
        sheet1[f'A{i+9}']=i
        sheet1[f'A{i+9}'].alignment= Alignment(horizontal='center', vertical='center')
        for col in ['B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L','M','N','O']:
            sheet1[f'{col}{i+9}'].alignment= Alignment(horizontal='center', vertical='center')
        
    sheet1['A1'].alignment= Alignment(horizontal='center', vertical='center')
    sheet1['A2'].alignment= Alignment(horizontal='center', vertical='center')
    sheet1['A3'].alignment= Alignment(horizontal='center', vertical='center')
    sheet1['A4'].alignment= Alignment(horizontal='left', vertical='center')
    sheet1['A5'].alignment= Alignment(horizontal='left', vertical='center')
    sheet1['A6'].alignment= Alignment(horizontal='left', vertical='center')
    sheet1['A7'].alignment= Alignment(horizontal='left', vertical='center')

    
    
    column_range = ['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L','M','N','O']
    for i in range(7,10):
        for col in column_range :
            sheet1[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')
    
    for i in range(total_roll+10,total_roll+17):
        for col in ['B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L','M','N','O'] :
            sheet1[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')

    for i in range(1,total_roll+17):
        for col in column_range:
            sheet1[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
    
    for col in column_range:
        sheet1[f'{col}{total_roll+16}'].font=Font(bold=True)
          
    for i in range(1,10):
        for col in column_range:
          sheet1[f'{col}{i}'].font=Font(bold=True)
    
    sheet1[f'A{total_roll+10}'] = 'Count(Attempted)'
    sheet1[f'A{total_roll+11}'] = 'Average Marks'
    sheet1[f'A{total_roll+12}'] = f'Count(>={al_values_temp[3]}%)'
    sheet1[f'A{total_roll+13}'] = f'% Count(>={al_values_temp[3]}% w.r.t appeared)'
    sheet1[f'A{total_roll+14}'] = 'Count(>=Average Marks of class)'
    sheet1[f'A{total_roll+15}']= "% Count(>=Average Marks of class w.r.t appeared)"
    sheet1[f'A{total_roll+16}'] = f'AL(Based on >={al_values_temp[3]}% Count) (All COs)'
    
    sheet1[f'F{total_roll+19}'] = "COs"
    sheet1[f'F{total_roll+19}'].font=Font(bold=True)
    sheet1[f'G{total_roll+19}'] = "AL"
    sheet1[f'G{total_roll+19}'].font=Font(bold=True)
    sheet1[f'F{total_roll+20}'] = 'CO1'
    sheet1[f'F{total_roll+21}'] = 'CO2'
    sheet1[f'F{total_roll+22}'] = 'CO3'
    sheet1[f'F{total_roll+23}'] = 'CO4'
    sheet1[f'F{total_roll+24}'] = 'CO5'
    print(f"i am coscout:::{cosCount}")
    print(f"Type of cosCount: {type(cosCount)}")
    if(cosCount == 6):
        print("hello i am 6")
        sheet1[f'F{total_roll+25}'] = 'CO6'
    coTableEndMidsem = total_roll+26
    if(cosCount==5):
        coTableEndMidsem = total_roll+25

    print(f"coTableEndMidsem : {coTableEndMidsem} ")
    for i in range(total_roll+19,coTableEndMidsem):
        print(f"F{i}")
        print(f"G{i}")
        sheet1[f'F{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        sheet1[f'G{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        sheet1[f'F{i}'].alignment= Alignment(horizontal='center', vertical='center')
        sheet1[f'G{i}'].alignment= Alignment(horizontal='center', vertical='center')
        
    #<-----------End Semester Template------------->
    
    sheet2.column_dimensions['A'].width =42
    sheet2.column_dimensions['B'].width =22
    sheet2['A1']="Roll No."  
    sheet2['A1'].font=Font(bold=True) 
    sheet2['B1']="ESE(TH)"
    sheet2['B1'].font=Font(bold=True)     
    sheet2['B2']="ALL Qs"
    sheet2['B2'].font=Font(bold=True) 
    sheet2['B3']="CO"+basic_values_temp[8]+""
    sheet2['B3'].font=Font(bold=True) 
    
    for i in range(1,total_roll+1):
        sheet2[f'A{i+3}'] =i
       
    
    sheet2[f'A{total_roll+4}']="Count(Attempted)"
    sheet2[f'A{total_roll+5}']="Average Marks"
    
   
    sheet2[f'A{total_roll+6}']=f"Count(>={al_values_temp[4]}%)"
    
    sheet2[f'A{total_roll+7}']=f"% Count(>={al_values_temp[4]}% w.r.t appeared)"
    
    sheet2[f'A{total_roll+8}']="Count(>=Average Marks of class)"
    sheet2[f'A{total_roll+9}']="% Count(>=Average Marks of class w.r.t appeared)"
    
    sheet2[f'A{total_roll+10}']=f"AL(Based on >={al_values_temp[4]}% Count) (All COs)"
    sheet2[f'A{total_roll+10}'].font=Font(bold=True) 
    sheet2[f'B{total_roll+10}'].font=Font(bold=True)
    
    for i in range(1,total_roll+11):
        sheet2[f'A{i}'].alignment= Alignment(horizontal='center', vertical='center')     
        sheet2[f'A{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        sheet2[f'B{i}'].alignment= Alignment(horizontal='center', vertical='center')     
        sheet2[f'B{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        if i>=total_roll+4 :
            sheet2[f'A{i}'].alignment= Alignment(horizontal='left', vertical='center') 
     
    sheet2[f'A{total_roll+13}'] = "COs"
    sheet2[f'A{total_roll+13}'].font=Font(bold=True)
    sheet2[f'B{total_roll+13}'] = "AL"
    sheet2[f'B{total_roll+13}'].font=Font(bold=True)
    sheet2[f'A{total_roll+14}'] = 'CO1'
    sheet2[f'A{total_roll+15}'] = 'CO2'
    sheet2[f'A{total_roll+16}'] = 'CO3'
    sheet2[f'A{total_roll+17}'] = 'CO4'
    sheet2[f'A{total_roll+18}'] = 'CO5'
    if(cosCount == 6):
        sheet2[f'A{total_roll+19}'] = 'CO6' 
    
    coTableEndsem = total_roll+20
    if(cosCount == 5):
        coTableEndsem = total_roll+19
    for i in range(total_roll+13,coTableEndsem):
        print("EEEEEENNNNdsemmmmm")
        print(f"A{i}")
        print(f"B{i}")
        sheet2[f'A{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        sheet2[f'B{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
        sheet2[f'A{i}'].alignment= Alignment(horizontal='center', vertical='center')
        sheet2[f'B{i}'].alignment= Alignment(horizontal='center', vertical='center')
        
   
    def make_CA_Type_PPT(mysheet2,arr,al_value,cosCount):
        
        groupSize = int(arr[0])
        mysheet2.column_dimensions['B'].width = 40

        # Create headers
        mysheet2['A1'] = "Sr. No."
        mysheet2['B1'] = "Name of Students"
        mysheet2['C1'] = "Roll Number"
        mysheet2['D1'] = "Group Number"
        mysheet2['E1'] = "Topic(Presentation)"
        mysheet2['F1'] = "Marks"  # New "Marks" column
        mysheet2['G1'] = "COs (Enter COs as digits with comma e.g. 1,2)"

        # Make headers bold
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            mysheet2[f'{col}1'].font = Font(bold=True)

        # Fill in Sr. No.
        for i in range(1, total_roll + 1):
            mysheet2[f'A{i + 1}'] = i

        # Center align and apply borders to header cells
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            cell = mysheet2[f'{col}1']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=Side(style='thin', color='000000'),
                                right=Side(style='thin', color='000000'),
                                left=Side(style='thin', color='000000'),
                                bottom=Side(style='thin', color='000000'))

        # Calculate the number of complete groups and handle any remaining students
        num_groups = total_roll // groupSize
        remaining_students = total_roll % groupSize

        # Process each group
        group_number = 1
        row_counter = 2
        for _ in range(num_groups):
            end_row = row_counter + groupSize - 1

            # Fill in the group number
            for j in range(row_counter, end_row + 1):
                mysheet2[f'D{j}'] = group_number

            # Merge cells in the "COs" column for the group
            mysheet2.merge_cells(f'G{row_counter}:G{end_row}')
            
            # Apply alignment to the merged cell
            cell = mysheet2[f'G{row_counter}']
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply borders to the entire group (Sr. No., Name, Roll Number, Group Number, Topic, Marks)
            for row in range(row_counter, end_row + 1):
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    mysheet2[f'{col}{row}'].border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000') if row == row_counter else Side(style=None),
                        bottom=Side(style='thin', color='000000') if row == end_row else Side(style=None)
                    )

            group_number += 1
            row_counter += groupSize

        # Handle the remaining students
        if remaining_students > 0:
            end_row = row_counter + groupSize - 1

            # Fill in the group number for remaining students
            for j in range(row_counter, row_counter + remaining_students):
                mysheet2[f'D{j}'] = group_number

            # Add empty cells for the remaining part of the last group
            for j in range(row_counter + remaining_students, end_row + 1):
                mysheet2[f'A{j}'] = ""
                mysheet2[f'B{j}'] = ""
                mysheet2[f'C{j}'] = ""
                mysheet2[f'D{j}'] = group_number
                mysheet2[f'E{j}'] = ""
                mysheet2[f'F{j}'] = ""

            # Merge cells in the "COs" column for the last group
            mysheet2.merge_cells(f'G{row_counter}:G{end_row}')
            
            # Apply alignment to the merged cell
            cell = mysheet2[f'G{row_counter}']
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply borders to the entire group
            for row in range(row_counter, end_row + 1):
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    mysheet2[f'{col}{row}'].border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000') if row == row_counter else Side(style=None),
                        bottom=Side(style='thin', color='000000') if row == end_row else Side(style=None)
                    )
                    
                    
    def make_CA_Type_NPTEL(mysheet2,arr,al_value,cosCount):
        mysheet2.column_dimensions['A'].width = 42
        mysheet2.column_dimensions['B'].width = 16
        mysheet2.merge_cells("A1:B1")
        mysheet2.merge_cells("A2:B2")
        mysheet2.merge_cells("A3:B3")
        mysheet2.merge_cells("A4:B4")
        mysheet2['A1']="Type :NPTEL Total Questions :"+str(len(arr))
        mysheet2['A1'].font=Font(bold=True)
        mysheet2['A3']="CA Marksheet"
        mysheet2['A3'].font=Font(bold=True)
        mysheet2['A5']="Roll No."
        mysheet2['A5'].font=Font(bold=True)
        mysheet2['B5']="CA1"
        mysheet2['B5'].font=Font(bold=True)
        mysheet2['B6']="CO"+arr[0]+""
        mysheet2['B6'].font=Font(bold=True)
        
        for i in range(1 ,total_roll+1):
            mysheet2[f'A{i+6}']=i
        
        mysheet2[f'A{total_roll+7}']="Count(Attempted)"
        mysheet2[f'A{total_roll+8}']="Average Marks"
        
    
        mysheet2[f'A{total_roll+9}']=f"Count(>={al_value}%)"
        
        
        mysheet2[f'A{total_roll+10}']=f"% Count(>={al_value}% w.r.t appeared)"
        
        mysheet2[f'A{total_roll+11}']="Count(>=Average Marks of class)"
        mysheet2[f'A{total_roll+12}']="% Count(>=Average Marks of class w.r.t appeared)"
        
        mysheet2[f'A{total_roll+13}']=f"AL(Based on >={al_value}% Count) (All COs)"
        mysheet2[f'A{total_roll+13}'].font=Font(bold=True)
        
        for i in range(1,total_roll+14):
            mysheet2[f'A{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet2[f'A{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            mysheet2[f'B{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet2[f'B{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
           
            if i>=total_roll+7 :
                mysheet2[f'A{i}'].alignment= Alignment(horizontal='left', vertical='center') 
        
        coTableEndnptel = total_roll+23
        if(cosCount==5):
            coTableEndnptel = total_roll+22
        for i in range(total_roll+16,coTableEndnptel):
            mysheet2[f'B{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet2[f'B{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            mysheet2[f'C{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet2[f'C{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            
        mysheet2[f'B{total_roll+16}'] = "COs"
        mysheet2[f'B{total_roll+16}'].font=Font(bold=True)
        mysheet2[f'C{total_roll+16}'] = "AL"
        mysheet2[f'C{total_roll+16}'].font=Font(bold=True)
        mysheet2[f'B{total_roll+17}'] = 'CO1'
        mysheet2[f'B{total_roll+18}'] = 'CO2'
        mysheet2[f'B{total_roll+19}'] = 'CO3'
        mysheet2[f'B{total_roll+20}'] = 'CO4'
        mysheet2[f'B{total_roll+21}'] = 'CO5'
        if(cosCount == 6):
            mysheet2[f'B{total_roll+22}'] = 'CO6' 
    
    def make_CA_Type_Quiz(mysheet,ca_array,al_value,cosCount):
        temp=len(ca_array)
        # mysheet=sheet4
        
        mysheet.column_dimensions['B'].width =42
        
        mysheet['A2']="Roll No."
        mysheet['B2']="Name"
        
        if temp==1 :
            mysheet['C2']="Q1"
            mysheet['C3']="CO"+ca_array[0]+""
            mysheet.merge_cells("A1:C1")
        
            myArr=['A','B','C']
            
        
        elif temp==2:
            mysheet['C2']="Q1"
            mysheet['D2']="Q2"
            
            mysheet['C3']="CO"+ca_array[0]+""
            mysheet['D3']="CO"+ca_array[1]+""
            mysheet.merge_cells("A1:D1")

            myArr=['A','B', 'C', 'D']
            
        elif temp==3:
            mysheet['C2']="Q1"
            mysheet['D2']="Q2"
            mysheet['E2']="Q3"
            
            mysheet['C3']="CO"+ca_array[0]+""
            mysheet['D3']="CO"+ca_array[1]+""
            mysheet['E3']="CO"+ca_array[2]+""
            mysheet.merge_cells("A1:E1")
            myArr=['A','B', 'C', 'D', 'E']
            
        elif temp==4:
            mysheet['C2']="Q1"
            mysheet['D2']="Q2"
            mysheet['E2']="Q3"
            mysheet['F2']="Q4"
            
            mysheet['C3']="CO"+ca_array[0]+""
            mysheet['D3']="CO"+ca_array[1]+""
            mysheet['E3']="CO"+ca_array[2]+""
            mysheet['F3']="CO"+ca_array[3]+""
            mysheet.merge_cells("A1:F1")
            myArr=['A','B', 'C', 'D', 'E', 'F']
            
        elif temp==5:
            mysheet['C2']="Q1"
            mysheet['D2']="Q2"
            mysheet['E2']="Q3"
            mysheet['F2']="Q4"
            mysheet['G2']="Q5"
            
            mysheet['C3']="CO"+ca_array[0]+""
            mysheet['D3']="CO"+ca_array[1]+""
            mysheet['E3']="CO"+ca_array[2]+""
            mysheet['F3']="CO"+ca_array[3]+""
            mysheet['G3']="CO"+ca_array[4]+""
            mysheet.merge_cells("A1:G1")
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G']
            
        elif temp==6:
            mysheet['C2']="Q1"
            mysheet['D2']="Q2"
            mysheet['E2']="Q3"
            mysheet['F2']="Q4"
            mysheet['G2']="Q5"
            mysheet['H2']="Q6"
            
            mysheet['C3']="CO"+ca_array[0]+""
            mysheet['D3']="CO"+ca_array[1]+""
            mysheet['E3']="CO"+ca_array[2]+""
            mysheet['F3']="CO"+ca_array[3]+""
            mysheet['G3']="CO"+ca_array[4]+""
            mysheet['H3']="CO"+ca_array[5]+""
            mysheet.merge_cells("A1:H1")
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G','H']
            
        elif temp==7:
            mysheet['C2']="Q1"
            mysheet['D2']="Q2"
            mysheet['E2']="Q3"
            mysheet['F2']="Q4"
            mysheet['G2']="Q5"
            mysheet['H2']="Q6"
            mysheet['I2']="Q7"
            
            mysheet['C3']="CO"+ca_array[0]+""
            mysheet['D3']="CO"+ca_array[1]+""
            mysheet['E3']="CO"+ca_array[2]+""
            mysheet['F3']="CO"+ca_array[3]+""
            mysheet['G3']="CO"+ca_array[4]+""
            mysheet['H3']="CO"+ca_array[5]+""
            mysheet['I3']="CO"+ca_array[6]+""
            mysheet.merge_cells("A1:I1")
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G','H','I']
            
        elif temp==8:
            mysheet['C2']="Q1"
            mysheet['D2']="Q2"
            mysheet['E2']="Q3"
            mysheet['F2']="Q4"
            mysheet['G2']="Q5"
            mysheet['H2']="Q6"
            mysheet['I2']="Q7"
            mysheet['J2']="Q8"
            
            mysheet['C3']="CO"+ca_array[0]+""
            mysheet['D3']="CO"+ca_array[1]+""
            mysheet['E3']="CO"+ca_array[2]+""
            mysheet['F3']="CO"+ca_array[3]+""
            mysheet['G3']="CO"+ca_array[4]+""
            mysheet['H3']="CO"+ca_array[5]+""
            mysheet['I3']="CO"+ca_array[6]+""
            mysheet['J3']="CO"+ca_array[7]+""
            mysheet.merge_cells("A1:J1")
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J']
            
        elif temp==9:
            mysheet['C2']="Q1"
            mysheet['D2']="Q2"
            mysheet['E2']="Q3"
            mysheet['F2']="Q4"
            mysheet['G2']="Q5"
            mysheet['H2']="Q6"
            mysheet['I2']="Q7"
            mysheet['J2']="Q8"
            mysheet['K2']="Q9"
            
            mysheet['C3']="CO"+ca_array[0]+""
            mysheet['D3']="CO"+ca_array[1]+""
            mysheet['E3']="CO"+ca_array[2]+""
            mysheet['F3']="CO"+ca_array[3]+""
            mysheet['G3']="CO"+ca_array[4]+""
            mysheet['H3']="CO"+ca_array[5]+""
            mysheet['I3']="CO"+ca_array[6]+""
            mysheet['J3']="CO"+ca_array[7]+""
            mysheet['K3']="CO"+ca_array[8]+""
            mysheet.merge_cells("A1:K1")
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K']
            
        else:
            mysheet['C2']="Q1"
            mysheet['D2']="Q2"
            mysheet['E2']="Q3"
            mysheet['F2']="Q4"
            mysheet['G2']="Q5"
            mysheet['H2']="Q6"
            mysheet['I2']="Q7"
            mysheet['J2']="Q8"
            mysheet['K2']="Q9"
            mysheet['L2']="Q10"
            
            mysheet['C3']="CO"+ca_array[0]+""
            mysheet['D3']="CO"+ca_array[1]+""
            mysheet['E3']="CO"+ca_array[2]+""
            mysheet['F3']="CO"+ca_array[3]+""
            mysheet['G3']="CO"+ca_array[4]+""
            mysheet['H3']="CO"+ca_array[5]+""
            mysheet['I3']="CO"+ca_array[6]+""
            mysheet['J3']="CO"+ca_array[7]+""
            mysheet['K3']="CO"+ca_array[8]+""
            mysheet['L3']="CO"+ca_array[9]+"" 
            mysheet.merge_cells("A1:L1")
            myArr=['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L']
        
        mysheet['A1']="Type : Quiz      Total Questions :"+str(temp)
        mysheet['A1'].font=Font(bold=True)
        
        for i in range(1,4):
            for col in  ['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L'] :
                mysheet[f'{col}{i}'].font=Font(bold=True)
                
        mysheet.merge_cells('A3:B3')
        
        
        for i in range(1 ,total_roll+1):
            mysheet[f'A{i+3}']=i
        
        for i in range(total_roll+4,total_roll+12) :
            mysheet.merge_cells(f'A{i}:B{i}')
            
        for i in range(1,total_roll+11):
            for j,col in enumerate(myArr[:temp+2]) :
                mysheet[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
                mysheet[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
                if i>total_roll+3 :
                    mysheet[f'{col}{i}'].alignment= Alignment(horizontal='left', vertical='center')     
                
        for i in range(total_roll+4,total_roll+11):
            start_index=1
            for j, col in enumerate(myArr[start_index:temp+2],start=start_index+1) :
                mysheet[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
                mysheet[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            if i==total_roll+10:
                mysheet[f'{col}{i}'].font=Font(bold=True)        
                
    
        mysheet[f'A{total_roll+4}']="Count(Attempted)"       
        mysheet[f'A{total_roll+5}']="Average Marks"
        
    
        mysheet[f'A{total_roll+6}']=f"Count(>={al_value}%)"
        
        
        mysheet[f'A{total_roll+7}']=f"% Count(>={al_value}% w.r.t appeared)"

        mysheet[f'A{total_roll+8}']="Count(>=Average Marks of class)"
        mysheet[f'A{total_roll+9}']="% Count(>=Average Marks of class w.r.t appeared)"
        
        mysheet[f'A{total_roll+10}']=f"AL(Based on >={al_value}% Count) (All COs)"
        mysheet[f'A{total_roll+10}'].font=Font(bold=True)
        
        coTableEndQuiz = total_roll+20
        if(cosCount == 5):
            coTableEndQuiz = total_roll+19
        for i in range(total_roll+13,coTableEndQuiz):
            mysheet[f'C{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet[f'C{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            mysheet[f'D{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            mysheet[f'D{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            
        mysheet[f'C{total_roll+13}'] = "COs"
        mysheet[f'C{total_roll+13}'].font=Font(bold=True)
        mysheet[f'D{total_roll+13}'] = "AL"
        mysheet[f'D{total_roll+13}'].font=Font(bold=True)
        mysheet[f'C{total_roll+14}'] = 'CO1'
        mysheet[f'C{total_roll+15}'] = 'CO2'
        mysheet[f'C{total_roll+16}'] = 'CO3'
        mysheet[f'C{total_roll+17}'] = 'CO4'
        mysheet[f'C{total_roll+18}'] = 'CO5'
        if(cosCount==6):
            mysheet[f'C{total_roll+19}'] = 'CO6' 
    
    print(basic_values_temp[10])       
    if basic_values_temp[10]=="2":
        if basic_values_temp[11]=="Quiz":
            make_CA_Type_Quiz(sheet3,CA1_Co_arr_temp,al_values_temp[0],cosCount)
        elif basic_values_temp[11]=="NPTEL Course":
            make_CA_Type_NPTEL(sheet3,CA1_Co_arr_temp,al_values_temp[0],cosCount) 
        else :
            make_CA_Type_PPT(sheet3,CA1_Co_arr_temp,al_values_temp[0],cosCount)
            
        print("Hello",basic_values_temp[10])
        if basic_values_temp[12]=="Quiz":
            make_CA_Type_Quiz(sheet4,CA2_Co_arr_temp,al_values_temp[1],cosCount)
        elif basic_values_temp[12]=="NPTEL Course":
            make_CA_Type_NPTEL(sheet4,CA2_Co_arr_temp,al_values_temp[1],cosCount)
        else :
            make_CA_Type_PPT(sheet4,CA2_Co_arr_temp,al_values_temp[1],cosCount)
            
    elif basic_values_temp[10]=="3":
        if basic_values_temp[11]=="Quiz":
            make_CA_Type_Quiz(sheet3,CA1_Co_arr_temp,al_values_temp[0],cosCount)
        elif basic_values_temp[11]=="NPTEL Course":
            make_CA_Type_NPTEL(sheet3,CA1_Co_arr_temp,al_values_temp[0],cosCount)
        else :
            make_CA_Type_PPT(sheet3,CA1_Co_arr_temp,al_values_temp[0],cosCount)
            
        print("Hi",basic_values_temp[10]) 
        if basic_values_temp[12]=="Quiz":
            make_CA_Type_Quiz(sheet4,CA2_Co_arr_temp,al_values_temp[1],cosCount)
        elif basic_values_temp[12]=="NPTEL Course":
            make_CA_Type_NPTEL(sheet4,CA2_Co_arr_temp,al_values_temp[1],cosCount) 
        else:
            make_CA_Type_PPT(sheet4,CA2_Co_arr_temp,al_values_temp[1],cosCount)
                 
        if basic_values_temp[13]=="Quiz":
            make_CA_Type_Quiz(sheet7,CA3_Co_arr_temp,al_values_temp[2],cosCount)
        elif basic_values_temp[13]=="NPTEL Course":
            make_CA_Type_NPTEL(sheet7,CA3_Co_arr_temp,al_values_temp[2],cosCount) 
        else:
            make_CA_Type_PPT(sheet7,CA3_Co_arr_temp,al_values_temp[2],cosCount)
        print("Hi",basic_values_temp[13])   
        
        
    #<-----------Survey Template------------->
    sheet5.column_dimensions['B'].width =40
    sheet5.column_dimensions['C'].width =40
    sheet5.column_dimensions['F'].width =40
    
    sheet5['A1']="Sr. No."
    sheet5['B1']="Email Address"
    sheet5['C1']="Full name of Student"
    sheet5['D1']="Roll No."
    sheet5['E1']="Class"
    sheet5['F1']="Branch"
    sheet5['G1']="Q1"
    sheet5['H1']="Q2"
    sheet5['I1']="Q3"
    sheet5['J1']="Q4"
    sheet5['K1']="Q5"
    sheet5['L1']="Q6"
    
    for col in  ['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L'] :
            sheet5[f'{col}1'].font=Font(bold=True)
            
    for i in range(1 ,total_roll+1):
        sheet5[f'A{i+1}']=i
        sheet5[f'E{i+1}']=""+basic_values_temp[7]+""
        sheet5[f'F{i+1}']=""+basic_values_temp[1]+""
        
    for i in range(1,total_roll+2):
        for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H','I', 'J','K','L'] :
            sheet5[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            sheet5[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
    
    sheet5[f'F{total_roll+4}']= 'Total' 
    sheet5[f'F{total_roll+4}'].font=Font(bold=True) 
    sheet5[f'F{total_roll+5}']= 'SA + A Count'
    sheet5[f'F{total_roll+5}'].font=Font(bold=True)
    sheet5[f'F{total_roll+6}']= 'SA + A Percentage' 
    sheet5[f'F{total_roll+6}'].font=Font(bold=True)
    sheet5[f'F{total_roll+7}']= 'CO Mapped' 
    sheet5[f'F{total_roll+7}'].font=Font(bold=True)
    sheet5[f'F{total_roll+8}']= 'AL'
    
    for i in range(total_roll+4,total_roll+9):
        for col in ['F','G','H','I','J','K','L'] :
            sheet5[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
            sheet5[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))
            if col=="F":
                sheet5[f'F{i}'].alignment= Alignment(horizontal='left', vertical='center')     
    
            if i==total_roll+8:
                sheet5[f'{col}{i}'].font=Font(bold=True)
        
    sheet5[f'G{total_roll+7}']= 'CO1' 
    sheet5[f'H{total_roll+7}']= 'CO2' 
    sheet5[f'I{total_roll+7}']= 'CO3' 
    sheet5[f'J{total_roll+7}']= 'CO4' 
    sheet5[f'K{total_roll+7}']= 'CO5' 
    if(cosCount==6):
        sheet5[f'L{total_roll+7}']= 'CO6'
    
    
    
    #<-----------------------Attainment--------------------->
    if basic_values_temp[10]=="3":
        sheet6.column_dimensions['A'].width =16
        sheet6.column_dimensions['B'].width =25
        sheet6.column_dimensions['C'].width =25
        sheet6.column_dimensions['D'].width =25
        sheet6.column_dimensions['E'].width =25
        sheet6.column_dimensions['F'].width =25
        sheet6.column_dimensions['G'].width =34
        sheet6.column_dimensions['H'].width =25
        
        for i in range (1,9):
            sheet6.merge_cells(f"A{i}:H{i}")
            sheet6[f'A{i}'].font=Font(bold=True)
            
        
        for i in range (9,15):
            sheet6.merge_cells(f"B{i}:H{i}") 
            
        

        sheet6["A1"].value="Vivekanand Education Society's Institute of Technology"
        sheet6["A1"].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6["A2"].value="Department of "+basic_values_temp[1]+""
        sheet6["A2"].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6["A3"].value="Academic Year :"+basic_values_temp[5]+""
        sheet6["A3"].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6["A5"].value="  Subject : "+basic_values_temp[4]+"                                                                                                                                                                       Class : "+basic_values_temp[7]+""
        sheet6["A5"].alignment= Alignment(horizontal='left', vertical='center')     
        
        sheet6["A6"].value="  Subject Teacher :"+basic_values_temp[6]+"                                                                                                                                                                Semester : "+basic_values_temp[3]+""
        sheet6["A6"].alignment= Alignment(horizontal='left', vertical='center')     
        
        
        sheet6['A8']='Course Outcomes(COs): Upon successful completion of this course, students will be able to:'
        sheet6['A8'].font=Font(bold=True)
        sheet6["A8"].alignment= Alignment(horizontal='left', vertical='center')     
        
        sheet6['A9'] ='CO1'
        sheet6['A10']='CO2'
        sheet6['A11']='CO3'
        sheet6['A12']='CO4'
        sheet6['A13']='CO5'
        if(cosCount==6):
            sheet6['A14']='CO6'

        # print(coTextArray)

        
        sheet6['B9'].value =""+coTextArray[0]+""
        sheet6['B10'].value=""+coTextArray[1]+""
        sheet6['B11'].value=""+coTextArray[2]+""
        sheet6['B12'].value=""+coTextArray[3]+""
        sheet6['B13'].value=""+coTextArray[4]+""
        if(cosCount==6):
            sheet6['B14'].value=""+coTextArray[5]+""
        maxRange = 15
        if(cosCount==5):
            maxRange=14
        for i in range (9,maxRange):
            sheet6[f'A{i}'].alignment= Alignment(horizontal='center', vertical='center')
            sheet6[f'B{i}'].alignment= Alignment(horizontal='left', vertical='center')         
        
        for i in range(9,maxRange):
            for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H']:
                sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
                        
        sheet6.merge_cells("A15:H15")
        sheet6.merge_cells("A16:H16")
        
        sheet6['A16']='CO Rubrics Mapping'
        sheet6['A16'].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6.merge_cells("A17:H17")
        
        sheet6.merge_cells("A18:A19")  
        sheet6.merge_cells("G18:G19")
        
        sheet6.merge_cells("B18:F18")
        sheet6.merge_cells("B19:E19") 
        
        for i in range(16,21):
            for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H']:
                sheet6[f'{col}{i}'].font=Font(bold=True)
        
        maxRange2 = 27
        if(cosCount==5):
            maxRange2=26
        for i in range(18,maxRange2):
            for col in ['A','B', 'C', 'D', 'E', 'F','G']:
                sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000')) 
                sheet6[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
                    
        sheet6['A18']='Assessment'
        sheet6['B18']='Direct Assessment' 
        sheet6['G18']='Indirect Assessment' 
        
        sheet6['B19']='Internal Assessment' 
        sheet6['F19']='External Assessment' 
        
        sheet6['A20']="CO's"
        sheet6['B20']="Mid Term Test"
        sheet6['C20']="CA1"
        sheet6['D20']="CA2"
        sheet6['E20']="CA3"
        sheet6['F20']="ESE(TH)"
        sheet6['G20']="Course Exit Survey"
        
        sheet6['A21']='CO1'
        sheet6['A22']='CO2'
        sheet6['A23']='CO3'
        sheet6['A24']='CO4'
        sheet6['A25']='CO5'
        if(cosCount==6):
            sheet6['A26']='CO6'
        
        sheet6.merge_cells("A27:H27")
        sheet6.merge_cells("A28:H28")
        sheet6.merge_cells("A29:H29")
        
        sheet6['A28']='CO Attainment (Level)'
        sheet6['A28'].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6.merge_cells("A30:A31")
        sheet6.merge_cells("H30:H31")
        
        sheet6.merge_cells("B30:G30")
        sheet6.merge_cells("B31:E31")
        
        for i in range(28,33):
            for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H']:
                sheet6[f'{col}{i}'].font=Font(bold=True)
        
        maxRange3 = 39
        if(cosCount==5):
            maxRange3 = 38
        for i in range(30,maxRange3):
            for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H']:
                sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000')) 
                sheet6[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6['A30']='Assessment'
        sheet6['B30']='Direct Assessment'
        sheet6['H30']='Indirect Assessment'
        
        sheet6['B31']='Internal Assessment'
        sheet6['F31']='External Assessment'
        sheet6['G31']='Attainment Level'
        
        sheet6['A32']="CO's"
        sheet6['B32']="Mid Term Test"
        sheet6['C32']="CA1"
        sheet6['D32']="CA2"
        sheet6['E32']="CA3"
        sheet6['F32']="ESE(TH)"
        sheet6['G32']="70% (External) + 30% (Internal)"
        sheet6['H32']="Course Exit Survey"
        
        sheet6['A33']='CO1'
        sheet6['A34']='CO2'
        sheet6['A35']='CO3'
        sheet6['A36']='CO4'
        sheet6['A37']='CO5'
        if(cosCount==6):
            sheet6['A38']='CO6'
        
        sheet6.merge_cells("A39:H39")
        sheet6.merge_cells("A40:H40")
        sheet6.merge_cells("A41:H41")
        
        maxRange4 = 49
        if(cosCount == 5):
            maxRange4 = 48
        for i in range(42,maxRange4):
            for col in ['C', 'D', 'E']:
                sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000')) 
                sheet6[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6['A40']='Final CO Attainment'
        sheet6['A40'].font=Font(bold=True)
        sheet6['A40'].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6['C42']='Course Outcomes'
        sheet6['C42'].font=Font(bold=True)
        
        sheet6['D42']='Direct AL'
        sheet6['D42'].font=Font(bold=True)
        
        sheet6['E42']='Indirect AL'
        sheet6['E42'].font=Font(bold=True)
        
        sheet6['C43']='CO1'
        sheet6['C43'].font=Font(bold=True)
        
        sheet6['C44']='CO2'
        sheet6['C44'].font=Font(bold=True)
        
        sheet6['C45']='CO3'
        sheet6['C45'].font=Font(bold=True)
        
        sheet6['C46']='CO4'
        sheet6['C46'].font=Font(bold=True)
        
        sheet6['C47']='CO5'
        sheet6['C47'].font=Font(bold=True)
        if(cosCount==6):
            sheet6['C48']='CO6'
            sheet6['C48'].font=Font(bold=True)
        
    else:
        sheet6.column_dimensions['A'].width =16
        sheet6.column_dimensions['B'].width =25
        sheet6.column_dimensions['C'].width =25
        sheet6.column_dimensions['D'].width =25
        sheet6.column_dimensions['E'].width =25
        sheet6.column_dimensions['F'].width =34
        sheet6.column_dimensions['G'].width =25
        
        for i in range (1,9):
            sheet6.merge_cells(f"A{i}:H{i}")
            sheet6[f'A{i}'].font=Font(bold=True)
            
        
        for i in range (9,15):
            sheet6.merge_cells(f"B{i}:H{i}") 
            
        

        sheet6["A1"].value="Vivekanand Education Society's Institute of Technology"
        sheet6["A1"].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6["A2"].value="Department of "+basic_values_temp[1]+""
        sheet6["A2"].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6["A3"].value="Academic Year :"+basic_values_temp[5]+""
        sheet6["A3"].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6["A5"].value="  Subject : "+basic_values_temp[4]+"                                                                                                                                                                       Class : "+basic_values_temp[7]+""
        sheet6["A5"].alignment= Alignment(horizontal='left', vertical='center')     
        
        sheet6["A6"].value="  Subject Teacher :"+basic_values_temp[6]+"                                                                                                                                                                Semester : "+basic_values_temp[3]+""
        sheet6["A6"].alignment= Alignment(horizontal='left', vertical='center')     
        
        
        sheet6['A8']='Course Outcomes(COs): Upon successful completion of this course, students will be able to:'
        sheet6['A8'].font=Font(bold=True)
        sheet6["A8"].alignment= Alignment(horizontal='left', vertical='center')     
        
        sheet6['A9'] ='CO1'
        sheet6['A10']='CO2'
        sheet6['A11']='CO3'
        sheet6['A12']='CO4'
        sheet6['A13']='CO5'
        if(cosCount==6):
            sheet6['A14']='CO6'
        
        # print(coTextArray)

        sheet6['B9'].value =""+coTextArray[0]+""
        sheet6['B10'].value=""+coTextArray[1]+""
        sheet6['B11'].value=""+coTextArray[2]+""
        sheet6['B12'].value=""+coTextArray[3]+""
        sheet6['B13'].value=""+coTextArray[4]+""
        if(cosCount==6):
            sheet6['B14'].value=""+coTextArray[5]+""
        rangeMax = 15
        if(cosCount==5):
            rangeMax = 14
        
        for i in range (9,rangeMax):
            sheet6[f'A{i}'].alignment= Alignment(horizontal='center', vertical='center')
            sheet6[f'B{i}'].alignment= Alignment(horizontal='left', vertical='center')         
        
        for i in range(9,rangeMax):
            for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H']:
                sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
                        
        sheet6.merge_cells("A15:H15")
        sheet6.merge_cells("A16:H16")
        
        sheet6['A16']='CO Rubrics Mapping'
        sheet6['A16'].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6.merge_cells("A17:H17")
        
        sheet6.merge_cells("A18:A19")  
        sheet6.merge_cells("F18:F19")
        
        sheet6.merge_cells("B18:E18")
        sheet6.merge_cells("B19:D19") 
        
        for i in range(16,21):
            for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H']:
                sheet6[f'{col}{i}'].font=Font(bold=True)
        
        rangeMax2 = 27
        if(cosCount==5): rangeMax = 26
        for i in range(18,rangeMax2):
            for col in ['A','B', 'C', 'D', 'E', 'F']:
                sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000')) 
                sheet6[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
                    
        sheet6['A18']='Assessment'
        sheet6['B18']='Direct Assessment' 
        sheet6['F18']='Indirect Assessment' 
        
        sheet6['B19']='Internal Assessment' 
        sheet6['E19']='External Assessment' 
        
        sheet6['A20']="CO's"
        sheet6['B20']="Mid Term Test"
        sheet6['C20']="CA1"
        sheet6['D20']="CA2"
        sheet6['E20']="ESE(TH)"
        sheet6['F20']="Course Exit Survey"
        
        sheet6['A21']='CO1'
        sheet6['A22']='CO2'
        sheet6['A23']='CO3'
        sheet6['A24']='CO4'
        sheet6['A25']='CO5'
        if(cosCount==6):
            sheet6['A26']='CO6'
        
        sheet6.merge_cells("A27:H27")
        sheet6.merge_cells("A28:H28")
        sheet6.merge_cells("A29:H29")
        
        sheet6['A28']='CO Attainment (Level)'
        sheet6['A28'].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6.merge_cells("A30:A31")
        sheet6.merge_cells("G30:G31")
        
        sheet6.merge_cells("B30:F30")
        sheet6.merge_cells("B31:D31")
        
        for i in range(28,33):
            for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H']:
                sheet6[f'{col}{i}'].font=Font(bold=True)
        rangeMax3 = 39
        if(cosCount==5): rangeMax3=38
        for i in range(30,rangeMax3):
            for col in ['A','B', 'C', 'D', 'E', 'F', 'G']:
                sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000')) 
                sheet6[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6['A30']='Assessment'
        sheet6['B30']='Direct Assessment'
        sheet6['G30']='Indirect Assessment'
        
        sheet6['B31']='Internal Assessment'
        sheet6['E31']='External Assessment'
        sheet6['F31']='Attainment Level'
        
        sheet6['A32']="CO's"
        sheet6['B32']="Mid Term Test"
        sheet6['C32']="CA1"
        sheet6['D32']="CA2"
        sheet6['E32']="ESE(TH)"
        sheet6['F32']="70% (External) + 30% (Internal)"
        sheet6['G32']="Course Exit Survey"
        
        sheet6['A33']='CO1'
        sheet6['A34']='CO2'
        sheet6['A35']='CO3'
        sheet6['A36']='CO4'
        sheet6['A37']='CO5'
        if(cosCount==6):
            sheet6['A38']='CO6'
        
        sheet6.merge_cells("A39:H39")
        sheet6.merge_cells("A40:H40")
        sheet6.merge_cells("A41:H41")
        rangeMax4 = 49
        if(cosCount==5):
            rangeMax4 = 48
        for i in range(42,rangeMax4):
            for col in ['C', 'D', 'E']:
                sheet6[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000')) 
                sheet6[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6['A40']='Final CO Attainment'
        sheet6['A40'].font=Font(bold=True)
        sheet6['A40'].alignment= Alignment(horizontal='center', vertical='center')     
        
        sheet6['C42']='Course Outcomes'
        sheet6['C42'].font=Font(bold=True)
        
        sheet6['D42']='Direct AL'
        sheet6['D42'].font=Font(bold=True)
        
        sheet6['E42']='Indirect AL'
        sheet6['E42'].font=Font(bold=True)
        
        sheet6['C43']='CO1'
        sheet6['C43'].font=Font(bold=True)
        
        sheet6['C44']='CO2'
        sheet6['C44'].font=Font(bold=True)
        
        sheet6['C45']='CO3'
        sheet6['C45'].font=Font(bold=True)
        
        sheet6['C46']='CO4'
        sheet6['C46'].font=Font(bold=True)
        
        sheet6['C47']='CO5'
        sheet6['C47'].font=Font(bold=True)
        
        if(cosCount == 6):
            sheet6['C48']='CO6'
            sheet6['C48'].font=Font(bold=True)

    #--------------------PO---------------------------
    for i in range (1,9):
        sheet8.merge_cells(f"A{i}:O{i}")
        sheet8[f'A{i}'].font=Font(bold=True)
      
    
      
      
   #Delete it   
            
    # basic_values_temp=["ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC"]        

                
            

    sheet8["A1"].value="Vivekanand Education Society's Institute of Technology"
    sheet8["A1"].alignment= Alignment(horizontal='center', vertical='center')     

    sheet8["A2"].value="Department of "+basic_values_temp[1]+""
    sheet8["A2"].alignment= Alignment(horizontal='center', vertical='center')     

    sheet8["A3"].value="Academic Year :"+basic_values_temp[5]+""
    sheet8["A3"].alignment= Alignment(horizontal='center', vertical='center')     

    sheet8["A5"].value="  Subject : "+basic_values_temp[4]+"                                                                                                                                                                       Class : "+basic_values_temp[7]+""
    sheet8["A5"].alignment= Alignment(horizontal='left', vertical='center')     

    sheet8["A6"].value="  Subject Teacher :"+basic_values_temp[6]+"                                                                                                                                                                Semester : "+basic_values_temp[3]+""
    sheet8["A6"].alignment= Alignment(horizontal='left', vertical='center')     

    sheet8["A9"].value="Programme Outcomes(POs):"                                                                                                                       
    sheet8["A9"].alignment= Alignment(horizontal='left', vertical='center')     
    sheet8["A9"].font=Font(bold=True)
    sheet8.merge_cells("A9:O9")

    sheet8.merge_cells("A10:O10")
    sheet8["A10"].value="""PO1) Basic Engineering knowledge: An ability to apply the fundamental knowledge in mathematics, science and engineering to solve problems in Computer engineering.
    PO2) Problem Analysis: Identify, formulate, research literature and analyze computer engineering problems reaching substantiated conclusions using first principles of mathematics, natural sciences and computer engineering and sciences.
    PO3) Design/ Development of Solutions: Design solutions for complex computer engineering problems and design system components or processes that meet specified needs with appropriate consideration for public health and safety, cultural, societal and environmental considerations.
    PO4) Conduct investigations of complex engineering problems using research-based knowledge and research methods including design of experiments, analysis and interpretation of data and synthesis of information to provide valid conclusions
    PO5) Modern Tool Usage: Create, select and apply appropriate techniques, resources and modern computer engineering and IT tools including prediction and modeling to complex engineering activities with an understanding of the limitations. 
    PO6) The Engineer and Society: Apply reasoning informed by contextual knowledge to assess societal, health, safety, legal and cultural issues and the consequent responsibilities relevant to computer engineering practice.
    PO7) Environment and Sustainability: Understand the impact of professional computer engineering solutions in societal and environmental contexts and demonstrate knowledge of and need for sustainable development. 
    PO8) Ethics: Apply ethical principles and commit to professional ethics and responsibilities and norms of computer engineering practice.
    PO9) Individual and Team Work: Function effectively as an individual, and as a member or leader in diverse teams and in multidisciplinary settings. 
    PO10) Communication: Communicate effectively on complex engineering activities with the engineering community and with society at large, such as being able to comprehend and write effective reports and design documentation, make effective presentations and give and receive clear instructions 
    PO11) Project Management and Finance: Demonstrate knowledge and understanding of computer engineering and management principles and apply these to one's own work, as a member and leader in a team, to manage projects and in multidisciplinary environments.
    PO12) Life-long Learning: Recognize the need for and have the preparation and ability to engage in independent and lifelong learning in the broadest context of technological change.
    PSO1) Professional Skills - The ability to develop programs for computer based systems of varying complexity and domains using standard practices.
    PSO2) Successful Career - The ability to adopt skills, languages, environment and platforms for creating innovative carrier paths, being successful entrepreneurs or for pursuing higher studies."""    
    
    sheet8["A12"].value="CO - PO/PSO Mapping"                                                                                                                       
    sheet8["A12"].alignment= Alignment(horizontal='center', vertical='center')     
    sheet8["A12"].font=Font(bold=True)
    sheet8.merge_cells("A12:O12")
    

    
    sheet8.merge_cells("B14:M14")
    sheet8.merge_cells("N14:O14")
    sheet8.merge_cells("A14:A15")
    
    COrange = 22
    if(cosCount == 5):
        COrange = 21
    for i in range(14,COrange):
        for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
            sheet8[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
            sheet8[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')  
        

    for col in ['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
        sheet8[f'{col}15'].font=Font(bold=True)

    for i in range(14,COrange):
        sheet8[f'A{i}'].font=Font(bold=True)

    sheet8['A16']='CO1'
    sheet8['A17']='CO2'
    sheet8['A18']='CO3'
    sheet8['A19']='CO4'
    sheet8['A20']='CO5'
    if(cosCount==6):
        sheet8['A21']='CO6' 

    for col,i in zip(['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M'],range(1,13)):
        sheet8[f'{col}15']=f'PO{i}'
    
    sheet8['N15']='PSO1' 
    sheet8['O15']='PSO2'  

    sheet8['A14']='Course Outcomes'
    sheet8['A14'].alignment= Alignment(horizontal='center', vertical='center',wrap_text=True)  ######  With Wrap Text
    sheet8['B14']='Programme Outcomes' 
    sheet8['B14'].font=Font(bold=True)
    sheet8['N14']="PSOs" 
    sheet8['N14'].font=Font(bold=True)
        
        
        
        
    sheet8["A23"].value="Direct PO Attainment"                                                                                                                       
    sheet8["A23"].alignment= Alignment(horizontal='center', vertical='center')     
    sheet8["A23"].font=Font(bold=True)
    sheet8.merge_cells("A23:O23")
    

    
    sheet8.merge_cells("B25:M25")
    sheet8.merge_cells("N25:O25")
    sheet8.merge_cells("A25:A26")
    
    COrange = 33
    if(cosCount == 5):
        COrange = 32
    for i in range(25,COrange):
        for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
            sheet8[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
            sheet8[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')  
        

    for col in ['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
        sheet8[f'{col}26'].font=Font(bold=True)

    for i in range(25,33):
        sheet8[f'A{i}'].font=Font(bold=True)

    sheet8['A27']='CO1'
    sheet8['A28']='CO2'
    sheet8['A29']='CO3'
    sheet8['A30']='CO4'
    sheet8['A31']='CO5'
    if(cosCount==6):
        sheet8['A32']='CO6' 

    for col,i in zip(['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M'],range(1,13)):
        sheet8[f'{col}26']=f'PO{i}'
    
    sheet8['N26']='PSO1' 
    sheet8['O26']='PSO2'  

    sheet8['A25']='Course Outcomes(COs)'
    sheet8['A25'].alignment= Alignment(horizontal='center', vertical='center',wrap_text=True)  ######  With Wrap Text
    sheet8['B25']='Programme Outcomes(POs)' 
    sheet8['B25'].font=Font(bold=True)
    sheet8['N25']="PSOs"
    sheet8['N25'].font=Font(bold=True)



        
    sheet8["A34"].value="Direct PO Attainment (After Applying CO-PO Mapping)"                                                                                                                       
    sheet8["A34"].alignment= Alignment(horizontal='center', vertical='center')     
    sheet8["A34"].font=Font(bold=True)
    sheet8.merge_cells("A34:O34")
    

    
    sheet8.merge_cells("B36:M36")
    sheet8.merge_cells("N36:O36")
    sheet8.merge_cells("A36:A37")
    
    for i in range(36,45):
        for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
            sheet8[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
            sheet8[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')  
        

    for col in ['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
        sheet8[f'{col}37'].font=Font(bold=True)
        sheet8[f'{col}44'].font=Font(bold=True)

    for i in range(36,45):
        sheet8[f'A{i}'].font=Font(bold=True)

    sheet8['A38']='CO1'
    sheet8['A39']='CO2'
    sheet8['A40']='CO3'
    sheet8['A41']='CO4'
    sheet8['A42']='CO5'
    if(cosCount == 6):
        sheet8['A43']='CO6' 
    sheet8['A44']='Avg PO'

    for col,i in zip(['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M'],range(1,13)):
        sheet8[f'{col}37']=f'PO{i}'
    
    sheet8['N37']='PSO1' 
    sheet8['O37']='PSO2'  

    sheet8['A36']='Course Outcomes(COs)'
    sheet8['A36'].alignment= Alignment(horizontal='center', vertical='center',wrap_text=True)  ######  With Wrap Text
    sheet8['B36']='Programme Outcomes(POs)' 
    sheet8['B36'].font=Font(bold=True)
    sheet8['N36']="PSOs"
    sheet8['N36'].font=Font(bold=True)

    sheet8['A47']='AL' 
    sheet8['A47'].font=Font(bold=True)
    sheet8['B47']='%'
    sheet8['B47'].font=Font(bold=True)
    sheet8['A48']='1' 
    sheet8['A49']='2' 
    sheet8['A50']='3' 
    sheet8['B48']='40' 
    sheet8['B49']='60' 
    sheet8['B50']='100'

    for i in range(47,51):
            sheet8[f'A{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
            sheet8[f'A{i}'].alignment= Alignment(horizontal='center', vertical='center')  
            sheet8[f'B{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
            sheet8[f'B{i}'].alignment= Alignment(horizontal='center', vertical='center')  
            


    #----------------------CO Information - Optional----------------------------#

    sheet0.merge_cells('A1:C1')
    sheet0['A1'] = "Basic Information"
    sheet0['B3'] = al_values_temp[0]
    sheet0['B4'] = al_values_temp[1]
    sheet0['B5'] = al_values_temp[2]
    sheet0['B6'] = al_values_temp[3]
    sheet0['B7'] = al_values_temp[4]
    sheet0['B9'] = cosCount

    sheet0['A3'] = "TargetCA1Text"
    sheet0['A4'] = "TargetCA2Text"
    sheet0['A5'] = "TargetCA3Text"
    sheet0['A6'] = "TargetMidTermText"
    sheet0['A7'] = "TargetEndSemText"
    sheet0['A9'] = "No. of COs"
    
    selectedPath = filedialog.askdirectory()
    filePath = f'{selectedPath}/Template_{basic_values_temp[7]}_{basic_values_temp[4]}_{basic_values_temp[6]}_{basic_values_temp[5]}.xlsx'

    workbook.save(filePath)
    CTkMessagebox(message=f"Excel template downloaded successfully at {filePath}.",icon="check", option_1="OK")

    # EMAIL Part - need helps 

    def send_email(sender_email, sender_password, recipient_email, subject, body, file_path):
        try:
            # Create a multipart message
            message = MIMEMultipart()
            message['From'] = sender_email
            message['To'] = recipient_email
            message['Subject'] = subject

            # Attach the email body
            message.attach(MIMEText(body, 'plain'))

            # Attach the file
            with open(file_path, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())

            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename={os.path.basename(file_path)}"
            )
            message.attach(part)

            # Connect to the SMTP server and send the email
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login(sender_email, sender_password)
                server.sendmail(sender_email, recipient_email, message.as_string())
            print("Email sent successfully!")

        except Exception as e:
            print(f"Error sending email: {e}")

    # Main processing code
    def send_file():
        # Simulating file processing
        downloadCalculate = filePath

        # Notify the user
        print(f"Calculated excel sheet downloaded successfully at {downloadCalculate}.")

        # Input recipient email and other email details
        email_address = receiversEmail
        sender_email = "copoautomation@gmail.com"  # Replace with your email
        sender_password = "jbzs zfrc ibrg nelp"      # Replace with your email's app password
        subject = "Template Excel File"
        body = f"Please find the attached template Excel file - Template {basic_values_temp[7]}_{basic_values_temp[4]}_{basic_values_temp[6]}_{basic_values_temp[5]}.xlsx"

        # Send the file via email
        send_email(sender_email, sender_password, email_address, subject, body, downloadCalculate)

    # Call the function
    send_file()
   