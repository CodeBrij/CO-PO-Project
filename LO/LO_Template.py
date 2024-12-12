import openpyxl
from openpyxl.styles import Alignment,Font
from openpyxl.styles.borders import Border,Side
from openpyxl import Workbook
from tkinter import filedialog
from tkinter import messagebox
import tkinter as tk
import os
from pathlib import Path
import customtkinter as ctk
from CTkMessagebox import CTkMessagebox

workbook=Workbook()
sheet8=workbook.create_sheet(title='LO Attainment')
for i in range (1,9):
    sheet8.merge_cells(f"A{i}:O{i}")
    sheet8[f'A{i}'].font=Font(bold=True)
      
      
      
   #Delete it   
            
basic_values_temp=["ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC","ABC"]        

            
        

sheet8["A1"].value="Vivekanand Education Society's Institute of Technology"
sheet8["A1"].alignment= Alignment(horizontal='center', vertical='center')     

sheet8["A2"].value="Department of "+basic_values_temp[1]+""
sheet8["A2"].alignment= Alignment(horizontal='center', vertical='center')     

sheet8["A3"].value="Academic Year :"+basic_values_temp[5]+""
sheet8["A3"].alignment= Alignment(horizontal='center', vertical='center')     

sheet8["A5"].value="  Subject : "+basic_values_temp[4]+"                                                                                                                                                                       Class : "+basic_values_temp[7]+""
sheet8["A5"].alignment= Alignment(horizontal='left', vertical='center')     

sheet8["A6"].value="  Subject Teacher :"+basic_values_temp[6]+"                                                                                                                                                                Semester : "+basic_values_temp[3]+""
sheet8["A6"].alignment= Alignment(horizontal='left', vertical='center')     

sheet8["A9"].value="Programme Outcomes(POs):"                                                                                                                       
sheet8["A9"].alignment= Alignment(horizontal='left', vertical='center')     
sheet8["A9"].font=Font(bold=True)
sheet8.merge_cells("A9:O9")

sheet8.merge_cells("A10:O10")
sheet8["A10"].value="""PO1) Basic Engineering knowledge: An ability to apply the fundamental knowledge in mathematics, science and engineering to solve problems in Computer engineering.
PO2) Problem Analysis: Identify, formulate, research literature and analyze computer engineering problems reaching substantiated conclusions using first principles of mathematics, natural sciences and computer engineering and sciences.
PO3) Design/ Development of Solutions: Design solutions for complex computer engineering problems and design system components or processes that meet specified needs with appropriate consideration for public health and safety, cultural, societal and environmental considerations.
PO4) Conduct investigations of complex engineering problems using research-based knowledge and research methods including design of experiments, analysis and interpretation of data and synthesis of information to provide valid conclusions
PO5) Modern Tool Usage: Create, select and apply appropriate techniques, resources and modern computer engineering and IT tools including prediction and modeling to complex engineering activities with an understanding of the limitations. 
PO6) The Engineer and Society: Apply reasoning informed by contextual knowledge to assess societal, health, safety, legal and cultural issues and the consequent responsibilities relevant to computer engineering practice.
PO7) Environment and Sustainability: Understand the impact of professional computer engineering solutions in societal and environmental contexts and demonstrate knowledge of and need for sustainable development. 
PO8) Ethics: Apply ethical principles and commit to professional ethics and responsibilities and norms of computer engineering practice.
PO9) Individual and Team Work: Function effectively as an individual, and as a member or leader in diverse teams and in multidisciplinary settings. 
PO10) Communication: Communicate effectively on complex engineering activities with the engineering community and with society at large, such as being able to comprehend and write effective reports and design documentation, make effective presentations and give and receive clear instructions 
PO11) Project Management and Finance: Demonstrate knowledge and understanding of computer engineering and management principles and apply these to one's own work, as a member and leader in a team, to manage projects and in multidisciplinary environments.
PO12) Life-long Learning: Recognize the need for and have the preparation and ability to engage in independent and lifelong learning in the broadest context of technological change.
PSO1) Professional Skills - The ability to develop programs for computer based systems of varying complexity and domains using standard practices.
PSO2) Successful Career - The ability to adopt skills, languages, environment and platforms for creating innovative carrier paths, being successful entrepreneurs or for pursuing higher studies."""    
 
 
sheet8["A12"].value="CO - PO/PSO Mapping"                                                                                                                       
sheet8["A12"].alignment= Alignment(horizontal='center', vertical='center')     
sheet8["A12"].font=Font(bold=True)
sheet8.merge_cells("A12:O12")
 

 
sheet8.merge_cells("B14:M14")
sheet8.merge_cells("N14:O14")
sheet8.merge_cells("A14:A15")
 
for i in range(14,22):
    for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
        sheet8[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
        sheet8[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')  
       

for col in ['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
    sheet8[f'{col}15'].font=Font(bold=True)

for i in range(14,22):
    sheet8[f'A{i}'].font=Font(bold=True)

sheet8['A16']='CO1'
sheet8['A17']='CO2'
sheet8['A18']='CO3'
sheet8['A19']='CO4'
sheet8['A20']='CO5'
sheet8['A21']='CO6' 

for col,i in zip(['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M'],range(1,13)):
    sheet8[f'{col}15']=f'PO{i}'
 
sheet8['N15']='PSO1' 
sheet8['O15']='PSO2'  

sheet8['A14']='Course Outcomes'
sheet8['A14'].alignment= Alignment(horizontal='center', vertical='center',wrap_text=True)  ######  With Wrap Text
sheet8['B14']='Programme Outcomes' 
sheet8['B14'].font=Font(bold=True)
sheet8['N14']="PSOs" 
sheet8['N14'].font=Font(bold=True)
     
      
      
      
sheet8["A23"].value="Direct PO Attainment"                                                                                                                       
sheet8["A23"].alignment= Alignment(horizontal='center', vertical='center')     
sheet8["A23"].font=Font(bold=True)
sheet8.merge_cells("A23:O23")
 

 
sheet8.merge_cells("B25:M25")
sheet8.merge_cells("N25:O25")
sheet8.merge_cells("A25:A26")
 
for i in range(25,33):
    for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
        sheet8[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
        sheet8[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')  
       

for col in ['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
    sheet8[f'{col}26'].font=Font(bold=True)

for i in range(25,33):
    sheet8[f'A{i}'].font=Font(bold=True)

sheet8['A27']='CO1'
sheet8['A28']='CO2'
sheet8['A29']='CO3'
sheet8['A30']='CO4'
sheet8['A31']='CO5'
sheet8['A32']='CO6' 

for col,i in zip(['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M'],range(1,13)):
    sheet8[f'{col}26']=f'PO{i}'
 
sheet8['N26']='PSO1' 
sheet8['O26']='PSO2'  

sheet8['A25']='Course Outcomes(COs)'
sheet8['A25'].alignment= Alignment(horizontal='center', vertical='center',wrap_text=True)  ######  With Wrap Text
sheet8['B25']='Programme Outcomes(POs)' 
sheet8['B25'].font=Font(bold=True)
sheet8['N25']="PSOs"
sheet8['N25'].font=Font(bold=True)



      
sheet8["A34"].value="Direct PO Attainment (After Applying CO-PO Mapping)"                                                                                                                       
sheet8["A34"].alignment= Alignment(horizontal='center', vertical='center')     
sheet8["A34"].font=Font(bold=True)
sheet8.merge_cells("A34:O34")
 

 
sheet8.merge_cells("B36:M36")
sheet8.merge_cells("N36:O36")
sheet8.merge_cells("A36:A37")
 
for i in range(36,45):
    for col in ['A','B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
        sheet8[f'{col}{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
        sheet8[f'{col}{i}'].alignment= Alignment(horizontal='center', vertical='center')  
       

for col in ['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M','N','O']:
    sheet8[f'{col}37'].font=Font(bold=True)
    sheet8[f'{col}44'].font=Font(bold=True)

for i in range(36,45):
    sheet8[f'A{i}'].font=Font(bold=True)

sheet8['A38']='CO1'
sheet8['A39']='CO2'
sheet8['A40']='CO3'
sheet8['A41']='CO4'
sheet8['A42']='CO5'
sheet8['A43']='CO6' 
sheet8['A44']='Avg PO'

for col,i in zip(['B', 'C', 'D', 'E', 'F', 'G','H','I','J','K','L','M'],range(1,13)):
    sheet8[f'{col}37']=f'PO{i}'
 
sheet8['N37']='PSO1' 
sheet8['O37']='PSO2'  

sheet8['A36']='Course Outcomes(COs)'
sheet8['A36'].alignment= Alignment(horizontal='center', vertical='center',wrap_text=True)  ######  With Wrap Text
sheet8['B36']='Programme Outcomes(POs)' 
sheet8['B36'].font=Font(bold=True)
sheet8['N36']="PSOs"
sheet8['N36'].font=Font(bold=True)

sheet8['A47']='AL' 
sheet8['A47'].font=Font(bold=True)
sheet8['B47']='%'
sheet8['B47'].font=Font(bold=True)
sheet8['A48']='1' 
sheet8['A49']='2' 
sheet8['A50']='3' 
sheet8['B48']='40' 
sheet8['B49']='60' 
sheet8['B50']='100'

for i in range(47,51):
        sheet8[f'A{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
        sheet8[f'A{i}'].alignment= Alignment(horizontal='center', vertical='center')  
        sheet8[f'B{i}'].border=Border(top=Side(style='thin',color='000000'),right=Side(style='thin',color='000000'),left=Side(style='thin',color='000000'),bottom=Side(style='thin',color='000000'))  
        sheet8[f'B{i}'].alignment= Alignment(horizontal='center', vertical='center')  
        

workbook.save("C:\\Users\\bhush\\Downloads\\LO.xlsx")