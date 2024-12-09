import openpyxl
import xlwings as xw

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Add some sample data to cells A1 to D1
ws['A1'] = 10
ws['B1'] = 20
ws['C1'] = 30
ws['D1'] = 40

# Save the workbook as .xlsx
excel_path = 'example.xlsx'
wb.save(excel_path)

# Open the .xlsx file with xlwings to add the macro and button
app = xw.App(visible=False)
wb_xlwings = app.books.open(excel_path)

# Add a VBA module with the macro code
macro_code = """
Sub AddValues()
    Range("E1").Value = Range("A1").Value + Range("B1").Value + Range("C1").Value + Range("D1").Value
End Sub
"""
vba_module = wb_xlwings.api.VBProject.VBComponents.Add(1)
vba_module.CodeModule.AddFromString(macro_code)

# Add a button to the worksheet
ws_xlwings = wb_xlwings.sheets[0]
button = ws_xlwings.api.OLEObjects().Add(ClassType="Forms.CommandButton.1", Link=False, DisplayAsIcon=False, Left=100, Top=50, Width=100, Height=30)
button_name = button.Name
button.Object.Caption = "Add A1 to D1"

# Ensure the button is correctly named and link the macro
button.Object.OnAction = f"'{wb_xlwings.name}!AddValues'"

# Save the workbook as .xlsm
xlsm_path = 'example_with_macro.xlsm'
wb_xlwings.save(xlsm_path)
wb_xlwings.close()
app.quit()
