import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook

##### ----  Lab Work   ---- ####
# Taking inputs
subject = input("Enter the subject name: ")
total_roll = int(input("Enter the total number of roll numbers: "))
total_exp = int(input("Enter the total number of experiments: "))
LabTarget = int(input("Enter the Lab Target: "))

# Create workbook and sheet
workbook = Workbook()
sheet1 = workbook.active
sheet1.title = "Lab"

# Setting the Subject Lab Work as the heading in A1
sheet1['A1'] = f"{subject} Lab Work"
sheet1['A1'].font = Font(size=14, bold=True)  # Make the heading bold and larger
sheet1['A1'].alignment = Alignment(horizontal='center')  # Center align the heading

# Merging the heading across columns for better formatting
sheet1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_exp+6)

# Setting the Lab Target value in A2
sheet1['A2'] = f"Target = {LabTarget}"

# Header row setup
sheet1['A3'] = "Roll No."
sheet1['B3'] = "Name"
sheet1['B2'] = f"Total Experiment = {total_exp}"
# Insert a new row after the 3rd row
sheet1.insert_rows(4)

# Set the "LO" in the new 4th row and merge columns A and B
sheet1['A5'] = "LO"
sheet1.merge_cells('A5:B5')
sheet1['A5'].font = Font(size=12, bold=True)
sheet1['A5'].alignment = Alignment(horizontal='center')

# Keeping other cells in this row blank (this is default, so no need to add values here)

# Adding experiment numbers horizontally in row 5 (since we inserted a row, everything shifts down by 1)
for i in range(total_exp):
    sheet1.cell(row=4, column=i+3, value=f"Exp {i+1}")  # Add "Exp 1", "Exp 2", etc., in row 5 across columns

# Adding columns for averages and assignments after the experiments
sheet1.cell(row=5, column=total_exp+3, value="Average(15)")
sheet1.cell(row=5, column=total_exp+4, value="Assignment 1")
sheet1.cell(row=5, column=total_exp+5, value="Assignment 2")
sheet1.cell(row=5, column=total_exp+6, value="Average Assignment(5)")

roll_end = 0
# Fill in roll numbers in the first column (column A) starting from row 6
for i in range(6, total_roll + 6):
    sheet1[f'A{i}'] = i - 5  # Roll number starts from 1
    roll_end = i-5


# Adjust column width for "Roll No." and "Name"
sheet1.column_dimensions['A'].width = 10  # Adjust Roll No. column width
sheet1.column_dimensions['B'].width = 25  # Adjust Name column width

sheet1[f'A{roll_end+2+5}'] = f'Count>={LabTarget}'
sheet1[f'A{roll_end+3+5}'] = f'%Count'
sheet1[f'A{roll_end+4+5}'] = f'AL'

# Adding borders to each cell
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Apply borders to the header row and roll number rows
for row in sheet1.iter_rows(min_row=3, max_row=total_roll+9, min_col=1, max_col=total_exp+6):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')



# Save the workbook
workbook.save("Lab_Records.xlsx")
