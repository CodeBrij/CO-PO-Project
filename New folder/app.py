import openpyxl
from openpyxl import load_workbook

wb = load_workbook('test.xlsx')

surveySheet = wb['Sheet1']

char_Sur = 'A'
surRow = 1

for row in range (2,100):
    if(surveySheet[f'{char_Sur+str(row)}'].value is None):
            surRow = row
            break

columnLetter = []

for col in range (5,50):
    colLetter = openpyxl.utils.get_column_letter(col)
    if(surveySheet[f'{colLetter+'1'}'].value is not None):
        columnLetter.append(colLetter)
    else:
        break

surveySheet[f'{'D'+str(surRow)}'] = 'Total'
surveySheet[f'{'D'+str(surRow+1)}'] = 'SA + A Count'
surveySheet[f'{'D'+str(surRow+2)}'] = 'SA + A Percentage'
surveySheet[f'{'D'+str(surRow+3)}'] = 'CO Mapped'
surveySheet[f'{'D'+str(surRow+4)}'] = 'AL'

for index in range (0,len(columnLetter)):
    surveySheet[f'{columnLetter[index]+str(surRow)}'] = f'=COUNT({columnLetter[index]+'2'}:{columnLetter[index]+str(surRow-1)})'
    surveySheet[f'{columnLetter[index]+str(surRow+1)}'] = f'=COUNTIF({columnLetter[index]+'2'}:{columnLetter[index]+str(surRow-1)}, ">=4")'
    surveySheet[f'{columnLetter[index]+str(surRow+2)}'] = f'=ROUND(({columnLetter[index]+str(surRow+1)}/{columnLetter[index]+str(surRow)}*100), 1)'
    surveySheet[f'{columnLetter[index]+str(surRow+3)}'] = 'CO'+f'{str(index+1)}'
    surveySheet[f'{columnLetter[index]+str(surRow+4)}'] = f'=IF({columnLetter[index]+str(surRow+2)}<60,1,IF(AND({columnLetter[index]+str(surRow+2)}>59,{columnLetter[index]+str(surRow+2)}<70),2,IF(AND({columnLetter[index]+str(surRow+2)}>69,{columnLetter[index]+str(surRow+2)}<80),3,4)))'


wb.save('test.xlsx')
