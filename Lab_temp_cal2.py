# import re
# from openpyxl.styles import *
# from openpyxl import *

# workbook = load_workbook('IoE_Lab.xlsx')
# sheet = workbook['Lab']
# rollCount = 0
# for i in 200:
#     if(sheet[f'A{i}'].value == "AL"):
#         rollCount = i

# target = sheet['A2'].value
# match = re.search(r'\d+', target)
# if match:
#     targetvalue = int(match.group())

# calStart = rollCount-2
# rollCount = rollCount-7
# startCount = 5
# startCol = 'E'
# endCol = 999
# endRow = rollCount+4
# for j in 26:
#     if(sheet.cell(row=3, column=j).value == None):
#         endCol='A'+j
# j='A'
# for j in endCol: 
#     sheet[f'{startCol}{calStart}'] = f"=COUNTIF({startCol}{startCount}:{startCol}{72}>={5*targetvalue/100})"
#     sheet[f'{startCol}{calStart+1}'] = f"=ROUND(({startCol}{calStart+1}/{targetvalue}*100),1)"
#     sheet[f'{startCol}{calStart+2}'] = f"=IF({startCol}{calStart+1}<60,1,IF(AND({startCol}{calStart+1}>59,{startCol}{calStart+1}<70),2,IF(AND({startCol}{calStart+1}>69,{startCol}{calStart+1}<80),3,4)))"
#     startCol = startCol + 1
    
import re
from openpyxl.styles import *
from openpyxl import *

# Load the workbook and select the sheet
workbook = load_workbook('IoE_Lab.xlsx')
sheet = workbook['Lab']

# Find the roll count by identifying the row where "AL" is located
rollCount = 0
for i in range(1, 200):  # Corrected to iterate over a range of rows
    if sheet[f'A{i}'].value == "AL":
        rollCount = i
        break  # Stop once we find "AL"

# Get the target value from cell A2
target = sheet['A2'].value
match = re.search(r'\d+', target)
if match:
    targetvalue = int(match.group())

# Set starting row and column for calculations
calStart = rollCount - 2  # Two rows above "AL"
rollCount = rollCount - 7  # Adjusting to where roll numbers are calculated
startCount = 5  # First row for student marks
startCol = 'E'  # First column for calculation (starting at column E)

# Find the last column for processing (where data ends)
endCol = None
for j in range(26, 999):  # Iterate through columns in the sheet
    if sheet.cell(row=3, column=j).value is None:  # Find the first empty column in row 3
        endCol = j - 1  # The last filled column
        break

# Step 1: Identify LOs in the 4th row (from column E onwards)
los = []
for col in range(ord(startCol), ord('A') + endCol + 1):  # Iterate through columns from E to endCol
    current_col = chr(col)  # Convert column number to letter
    lo_value = sheet[f'{current_col}4'].value  # Get the LO value in row 4
    if lo_value:
        los.append(lo_value)

# Step 2: Find the row for the new table (at the bottom of the sheet)
new_table_start_row = rollCount + 10  # Start the new table 10 rows after rollCount

# Step 3: Populate the LOs column and calculate AL
sheet[f'A{new_table_start_row}'] = "LOs"
sheet[f'B{new_table_start_row}'] = "AL"

for i, lo_value in enumerate(los):
    row = new_table_start_row + i + 1  # Row for each LO and AL

    # Insert LO value in column A
    sheet[f'A{row}'] = lo_value

    # Insert average AL value (using the AL row)
    al_formula = f"=AVERAGE({chr(ord(startCol) + i)}{rollCount})"  # Average from AL row
    sheet[f'B{row}'] = al_formula

# Save the workbook after adding the new table
workbook.save('IoE_Lab_Updated_with_LO_AL_Table.xlsx')
