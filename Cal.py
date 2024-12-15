import openpyxl
from openpyxl.styles import Alignment,Font
from openpyxl.styles.borders import Border,Side
from openpyxl import Workbook
import re
import os
from tkinter import filedialog
from tkinter import messagebox
import customtkinter as ctk
from CTkMessagebox import CTkMessagebox
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

def cal_sheet(file_name, al_values_temp):
    workbook=openpyxl.load_workbook(file_name)
    print(file_name)
    file_name_only = os.path.basename(file_name)
    file_name_only = os.path.splitext(file_name_only)[0]
    file_name_only = file_name_only.replace("Template","")

    sheet = workbook['Midsem']
    cosCount=6
    for row in range(1,10):
        if(str(sheet[f'A{row}'].value)=="COs"):
            coRow = row
            break    

    for row in range(1,10):
        if (str(sheet[f'A{row}'].value)).startswith("Number of Students ="):
            total_roll=int((sheet[f'A{row}'].value).split('=')[-1].strip())
            break 

    # target=((sheet[f'A{coRow+total_roll+3}'].value).split('>=')[-1].strip()).replace('%)', '')
    # try:
    #     target = float(target)
    #     if target.is_integer():
    #         target = int(target)
    # except ValueError:  
    #     raise ValueError(f"Invalid target value: {target}")
    
    coRow =int(coRow) + 1  # Gives the start rowIndex for RollNo.
   
    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow}']=f'=COUNT({column_letter}{coRow}:{column_letter}{total_roll+coRow-1})'

    for row in range (coRow,total_roll+coRow):    
        sheet[f'H{row}']=f'=ROUND(SUM(B{row}:G{row}),0)'

    for row in range (coRow,total_roll+coRow):    
        sheet[f'K{row}']=f'=ROUND(SUM(I{row}:J{row}),0)'
        
    for row in range (coRow,total_roll+coRow):    
        sheet[f'N{row}']=f'=ROUND(SUM(L{row}:M{row}),0)'
            
    for row in range (coRow,total_roll+coRow):    
        sheet[f'O{row}']=f'=ROUND(SUM(H{row},K{row},N{row}),0)'

    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        formula = f'=IF(COUNT({column_letter}{coRow}:{column_letter}{total_roll+coRow-1})=0, "-", ROUND(AVERAGE({column_letter}{coRow}:{column_letter}{total_roll+coRow-1}), 0))'
        sheet[f'{column_letter}{total_roll+coRow+1}'] = formula
    
    for col in range (2,8):
        column_letter = openpyxl.utils.get_column_letter(col)
        target_cell = sheet[f'{column_letter}{total_roll+coRow+2}']
        if target_cell.value is None:  # Check if the cell is empty
            target_cell.value = f'=COUNTIF({column_letter}{coRow}:{column_letter}{total_roll+coRow-1},">={float(al_values_temp[3])/100 * 2}")'
    for col in range (8,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        target_cell = sheet[f'{column_letter}{total_roll+coRow+2}']
        if target_cell.value is None:  # Check if the cell is empty
            target_cell.value = f'=COUNTIF({column_letter}{coRow}:{column_letter}{total_roll+coRow-1},">={float(al_values_temp[3])/100 * 5}")'

    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow+3}'] = (
            f'=ROUND(IFERROR({column_letter}{total_roll+coRow+2}/{column_letter}{total_roll+coRow}, 0)*100, 1)'
        )


    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow+4}']=f'=COUNTIF({column_letter}{coRow}:{column_letter}{total_roll+coRow-1},">="&{column_letter}{total_roll+coRow+1})'

    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow+5}'] = (
            f'=ROUND(IFERROR({column_letter}{total_roll+coRow+4}/{column_letter}{total_roll+coRow}, 0)*100, 1)'
        )


    for col in range (2,16):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet[f'{column_letter}{total_roll+coRow+6}']=f'=IF({column_letter}{total_roll+coRow+3}<60,1,IF(AND({column_letter}{total_roll+coRow+3}>59,{column_letter}{total_roll+coRow+3}<70),2,IF(AND({column_letter}{total_roll+coRow+3}>69,{column_letter}{total_roll+coRow+3}<80),3,4)))'
        

    coTableRow = total_roll+coRow+9

    if(sheet[f'F{coTableRow+6}'].value is None):
        cosCount = 5

    print(f"COscount:::::{cosCount}")
    columns_with_1 = []
    columns_with_2 = []
    columns_with_3 = []
    columns_with_4 = []
    columns_with_5 = []
    columns_with_6 = [] 

        
    column_range = ['B', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'L', 'M']

    # Iterate through cells in the specified column range
    for column_letter in column_range:
        # Get the cell in row 2 corresponding to the column letter
        cell = sheet[f"{column_letter}{coRow-1}"]
        # Check if the cell has a value
        if cell.value :
            # values = [int(val.strip()) for val in str(cell.value).split(',') if val.strip().isdigit()]  #This for without CO like 1,2,3
            values = [int(val.strip()) for val in str(cell.value)[2:].split(',') if val.strip().isdigit()]   #This for with CO like CO1,2,3
            # Check if '1' is present in the list of values then 2 and 3,4and 5
            # print(values)
            if 1 in values:
                # Add the cell value to the list for calculation
                columns_with_1.append(column_letter)
            if 2 in values:
                # Add the cell value to the list for calculation
                columns_with_2.append(column_letter)
            if 3 in values:
                # Add the cell value to the list for calculation
                columns_with_3.append(column_letter)
            if 4 in values:
                # Add the cell value to the list for calculation
                columns_with_4.append(column_letter)
            if 5 in values:
                # Add the cell value to the list for calculation
                columns_with_5.append(column_letter)
            if 6 in values:
                columns_with_6.append(column_letter)

    # Calculate the average using Excel formula
    if columns_with_1:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_1])}),1)"
        sheet[f'G{coTableRow+1}'] = average_formula
    else:
        sheet[f'G{coTableRow+1}'] = '-'

        
    if columns_with_2:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_2])}),1)"
        sheet[f'G{coTableRow+2}'] = average_formula
    else:
        sheet[f'G{coTableRow+2}'] = '-'
    
    if columns_with_3:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_3])}),1)"
        sheet[f'G{coTableRow+3}'] = average_formula
    else:
        sheet[f'G{coTableRow+3}'] = '-'
        
    if columns_with_4:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_4])}),1)"
        sheet[f'G{coTableRow+4}'] = average_formula
    else:
        sheet[f'G{coTableRow+4}'] = '-'

    if columns_with_5:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_5])}),1)"
        sheet[f'G{coTableRow+5}'] = average_formula
    else:
        sheet[f'G{coTableRow+5}'] = '-'
    if(cosCount == 6):
        if columns_with_6:
            average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+coRow+6}' for col_letter in columns_with_6])}),1)"
            sheet[f'G{coTableRow+6}'] = average_formula
        else:
            sheet[f'G{coTableRow+6}'] = '-'
    
    map_midsem_co_arr=[f"={sheet.title}!G{coTableRow+1}",f"={sheet.title}!G{coTableRow+2}",f"={sheet.title}!G{coTableRow+3}",f"={sheet.title}!G{coTableRow+4}",f"={sheet.title}!G{coTableRow+5}",f"={sheet.title}!G{coTableRow+6}"]
   #<----------------------End Sem------------------->
   
    sheet1 = workbook['Endsem']

    endCol='B'
    
    sheet1[f'B{total_roll+4}'] = f'=COUNT({endCol}4:{endCol}{total_roll+3})'
    sheet1[f'B{total_roll+5}'] = f'=ROUND(AVERAGE({endCol}4:{endCol}{total_roll+3}), 0)'
    target_cell = sheet1[f'B{total_roll+6}']
    if target_cell.value is None:  # Check if the cell is empty
        target_cell.value = f'=COUNTIF({endCol}4:{endCol}{total_roll+3}, ">={float(al_values_temp[4]) / 100 * 60}")'
    sheet1[f'B{total_roll+7}'] = f'=ROUND({sheet1[f"B{str(total_roll+6)}"].coordinate} / {sheet1[f"B{str(total_roll+4)}"].coordinate} * 100, 1)'
    sheet1[f'B{total_roll+8}'] = f'=COUNTIF({endCol}4:{endCol}{total_roll+3}, ">="&{endCol}{total_roll+5})'
    sheet1[f'B{total_roll+9}'] = f'=ROUND({sheet1[f"B{str(total_roll+8)}"].coordinate} / {sheet1[f"B{str(total_roll+4)}"].coordinate} * 100, 1)'
    sheet1[f'B{total_roll+10}'] = f'=IF({sheet1[f"B{str(total_roll+7)}"].coordinate}<60, 1, IF(AND({sheet1[f"B{str(total_roll+7)}"].coordinate}>59, {sheet1[f"B{str(total_roll+7)}"].coordinate}<70), 2, IF(AND({sheet1[f"B{str(total_roll+7)}"].coordinate}>69, {sheet1[f"B{str(total_roll+7)}"].coordinate}<80), 3, 4)))'
    
        
    check = [int(val.strip()) for val in str(sheet1['B3'].value)[2:].split(',') if val.strip().isdigit()]
    
    if 1 in check:
        sheet1[f'B{total_roll+14}']=sheet1[f'B{total_roll+10}'].value
    else :
        sheet1[f'B{total_roll+14}']="-"
    
    if 2 in check:
        sheet1[f'B{total_roll+15}']=sheet1[f'B{total_roll+10}'].value
    else :
        sheet1[f'B{total_roll+15}']="-"
        
    if 3 in check:
        sheet1[f'B{total_roll+16}']=sheet1[f'B{total_roll+10}'].value
    else :
        sheet1[f'B{total_roll+16}']="-"
        
    if 4 in check:
        sheet1[f'B{total_roll+17}']=sheet1[f'B{total_roll+10}'].value
    else :
        sheet1[f'B{total_roll+17}']="-"
        
    if 5 in check:
        sheet1[f'B{total_roll+18}']=sheet1[f'B{total_roll+10}'].value
    else :
        sheet1[f'B{total_roll+18}']="-"
        
    if(cosCount == 6):
        if 6 in check:
            sheet1[f'B{total_roll+19}']=sheet1[f'B{total_roll+10}'].value
        else :
            sheet1[f'B{total_roll+19}']="-"
    
    map_endsem_co_arr=[f'={sheet1.title}!B{total_roll+14}',f'={sheet1.title}!B{total_roll+15}',f'={sheet1.title}!B{total_roll+16}',f'={sheet1.title}!B{total_roll+17}',f'={sheet1.title}!B{total_roll+18}',f'={sheet1.title}!B{total_roll+19}']
    #<----------------------CA------------------->
    # my_CA1_Co_arr=call_CA(workbook['CA1'])
    # my_CA2_Co_arr=call_CA(workbook['CA2'])
    # my_CA3_Co_arr=call_CA(workbook['CA3'])
    
    def call_CA(mySheet, al_value):
        al_value = al_value
        print(al_value)
        # mySheet=workbook[ca]
        
        match1 = re.search(r'Type :\s*([\w/]+)', mySheet['A1'].value)
        if match1:
            quiz_type = match1.group(1)
        
        match2 = re.search(r'Total Questions :(\d+)', mySheet['A1'].value)
        if match2:
            total_questions = int(match2.group(1))
        
        if match2:
            if total_questions==1 :
                myArr=['C']
            elif total_questions==2:
                myArr=['C', 'D']  
            elif total_questions==3:
                myArr=['C', 'D', 'E']  
            elif total_questions==4:
                myArr=['C', 'D', 'E', 'F']   
            elif total_questions==5: 
                myArr=['C', 'D', 'E', 'F', 'G']  
            elif total_questions==6:
                myArr=['C', 'D', 'E', 'F', 'G','H']  
            elif total_questions==7:
                myArr=['C', 'D', 'E', 'F', 'G','H','I']  
            elif total_questions==8:
                myArr=['C', 'D', 'E', 'F', 'G','H','I', 'J']   
            elif total_questions==9: 
                myArr=['C', 'D', 'E', 'F', 'G','H','I', 'J','K']  
            else:  
                myArr=['C', 'D', 'E', 'F', 'G','H','I', 'J','K','L']   
       
             
        if match1:
            if quiz_type=='Quiz':
                my_co_arr=cal_quiz(mySheet,myArr,al_value)
            else:
                my_co_arr=cal_NPTEL(mySheet,al_value)
        else:
            my_co_arr=cal_PPT(mySheet,al_value) 
        return my_co_arr
            
        
        
    def cal_quiz(newSheet,col_arr,al_value):
        for col in col_arr:
        
            newSheet[f'{col}{total_roll+4}'] = f'=COUNT({col}4:{col}{total_roll+3})'
            newSheet[f'{col}{total_roll+5}'] = f'=ROUND(AVERAGE({col}4:{col}{total_roll+3}), 0)'
            target_cell = newSheet[f'{col}{total_roll+6}']
            if target_cell.value is None:  # Check if the cell is empty
                target_cell.value = f'=COUNTIF({col}4:{col}{total_roll+3}, ">={float(al_value) / 100 * 2}")'
            newSheet[f'{col}{total_roll+7}'] = f'=ROUND({newSheet[f"{col}{total_roll+6}"].coordinate} / {newSheet[f"{col}{total_roll+4}"].coordinate} * 100, 1)'
            newSheet[f'{col}{total_roll+8}'] = f'=COUNTIF({col}3:{col}{total_roll+3}, ">="&{col}{total_roll+5})'
            newSheet[f'{col}{total_roll+9}'] = f'=ROUND({newSheet[f"{col}{total_roll+8}"].coordinate} / {newSheet[f"{col}{total_roll+4}"].coordinate} * 100, 1)'
            newSheet[f'{col}{total_roll+10}'] = f'=IF({newSheet[f"{col}{total_roll+7}"].coordinate}<60, 1, IF(AND({newSheet[f"{col}{total_roll+7}"].coordinate}>59, {newSheet[f"{col}{total_roll+7}"].coordinate}<70), 2, IF(AND({newSheet[f"{col}{total_roll+7}"].coordinate}>69, {newSheet[f"{col}{total_roll+7}"].coordinate}<80), 3, 4)))'
    
        columns_QZ_1 = []
        columns_QZ_2 = []
        columns_QZ_3 = []
        columns_QZ_4 = []
        columns_QZ_5 = []
        columns_QZ_6 = []
        
        for col in col_arr:
            cell = newSheet[f"{col}3"]
            if cell.value :
                value3 = [int(val.strip()) for val in str(cell.value)[2:].split(',') if val.strip().isdigit()]   #This for with CO like CO1,2,3
                if 1 in value3:
                    # Add the cell value to the list for calculation
                    columns_QZ_1.append(col)
                if 2 in value3:
                    # Add the cell value to the list for calculation
                    columns_QZ_2.append(col)
                if 3 in value3:
                    # Add the cell value to the list for calculation
                    columns_QZ_3.append(col)
                if 4 in value3:
                    # Add the cell value to the list for calculation
                    columns_QZ_4.append(col)
                if 5 in value3:
                    # Add the cell value to the list for calculation
                    columns_QZ_5.append(col)
                if 6 in value3:
                    columns_QZ_6.append(col)

        # Calculate the average using Excel formula
        if columns_QZ_1:
            average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+10}' for col_letter in columns_QZ_1])}),1)"
            newSheet[f'D{total_roll+14}'] = average_formula
        else:
            newSheet[f'D{total_roll+14}'] = '-'

            
        if columns_QZ_2:
            average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+10}' for col_letter in columns_QZ_2])}),1)"
            newSheet[f'D{total_roll+15}'] = average_formula
        else:
            newSheet[f'D{total_roll+15}'] = '-'
        
        if columns_QZ_3:
            average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+10}' for col_letter in columns_QZ_3])}),1)"
            newSheet[f'D{total_roll+16}'] = average_formula
        else:
            newSheet[f'D{total_roll+16}'] = '-'
            
        if columns_QZ_4:
            average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+10}' for col_letter in columns_QZ_4])}),1)"
            newSheet[f'D{total_roll+17}'] = average_formula
        else:
            newSheet[f'D{total_roll+17}'] = '-'

        if columns_QZ_5:
            average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+10}' for col_letter in columns_QZ_5])}),1)"
            newSheet[f'D{total_roll+18}'] = average_formula
        else:
            newSheet[f'D{total_roll+18}'] = '-'

        if(cosCount == 6):
            if columns_QZ_6:
                average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+10}' for col_letter in columns_QZ_6])}),1)"
                newSheet[f'D{total_roll+19}'] = average_formula
            else:
                newSheet[f'D{total_roll+19}'] = '-'
        
        map_quiz_co_arr=[f'={newSheet.title}!D{total_roll+14}',f'={newSheet.title}!D{total_roll+15}',f'={newSheet.title}!D{total_roll+16}',f'={newSheet.title}!D{total_roll+17}',f'={newSheet.title}!D{total_roll+18}',f'={newSheet.title}!D{total_roll+19}']
        
        return map_quiz_co_arr
        
        
    # #<----------------------Survey------------------->
    
    # sheet4=workbook['Survey']
    
    # for col in ['G','H','I','J','K','L']:
        
    #     sheet4[f'{col}{total_roll+4}'] = f'=COUNT({col}2:{col}{total_roll+1})'
    #     sheet4[f'{col}{total_roll+5}'] = f'=COUNTIF({col}2:{col}{total_roll+1}, ">=4")'
    #     sheet4[f'{col}{total_roll+6}'] = f'=ROUND(({col}{total_roll+5}/{col}{total_roll+4}*100), 1)'
    #     sheet4[f'{col}{total_roll+8}'] = f'=IF({col}{total_roll+6}<60,1,IF(AND({col}{total_roll+6}>59,{col}{total_roll+6}<70),2,IF(AND({col}{total_roll+6}>69,{col}{total_roll+6}<80),3,4)))'
        
    # map_survey_co_arr=[f'=Survey!G{total_roll+8}',f'=Survey!H{total_roll+8}',f'=Survey!I{total_roll+8}',f'=Survey!J{total_roll+8}',f'=Survey!K{total_roll+8}',f'=Survey!L{total_roll+8}']

    def cal_NPTEL(mySheet2, al_value):
        
        for col in ['B']:
            
            mySheet2[f'{col}{total_roll+7}'] = f'=COUNT({col}7:{col}{total_roll+6})'
            mySheet2[f'{col}{total_roll+8}'] = f'=ROUND(AVERAGE({col}7:{col}{total_roll+6}), 0)'
            mySheet2[f'{col}{total_roll+9}'] = f'=COUNTIF({col}7:{col}{total_roll+6}, ">={float(al_value) / 100 * 10}")'
            mySheet2[f'{col}{total_roll+10}'] = f'=ROUND({mySheet2[f"{col}{total_roll+9}"].coordinate} / {mySheet2[f"{col}{total_roll+7}"].coordinate} * 100, 1)'
            mySheet2[f'{col}{total_roll+11}'] = f'=COUNTIF({col}7:{col}{total_roll+6}, ">="&{col}{total_roll+8})'
            mySheet2[f'{col}{total_roll+12}'] = f'=ROUND({mySheet2[f"{col}{total_roll+11}"].coordinate} / {mySheet2[f"{col}{total_roll+7}"].coordinate} * 100, 1)'
            mySheet2[f'{col}{total_roll+13}'] = f'=IF({mySheet2[f"{col}{total_roll+10}"].coordinate}<60, 1, IF(AND({mySheet2[f"{col}{total_roll+10}"].coordinate}>59, {mySheet2[f"{col}{total_roll+10}"].coordinate}<70), 2, IF(AND({mySheet2[f"{col}{total_roll+10}"].coordinate}>69, {mySheet2[f"{col}{total_roll+10}"].coordinate}<80), 3, 4)))'
        
        
        columns_CA_1 = []
        columns_CA_2 = []
        columns_CA_3 = []
        columns_CA_4 = []
        columns_CA_5 = []
        columns_CA_6 = [] 
        
        for col in ['B']:
            cell = mySheet2[f"{col}6"]
            if cell.value :
                value2 = [int(val.strip()) for val in str(cell.value)[2:].split(',') if val.strip().isdigit()]   #This for with CO like CO1,2,3
                if 1 in value2:
                    # Add the cell value to the list for calculation
                    columns_CA_1.append(col)
                if 2 in value2:
                    # Add the cell value to the list for calculation
                    columns_CA_2.append(col)
                if 3 in value2:
                    # Add the cell value to the list for calculation
                    columns_CA_3.append(col)
                if 4 in value2:
                    # Add the cell value to the list for calculation
                    columns_CA_4.append(col)
                if 5 in value2:
                    # Add the cell value to the list for calculation
                    columns_CA_5.append(col)
                if 6 in value2:
                    columns_CA_6.append(col)

        # Calculate the average using Excel formula
        if columns_CA_1:
            average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+13}' for col_letter in columns_CA_1])}),1)"
            mySheet2[f'C{total_roll+17}'] = average_formula
        else:
            mySheet2[f'C{total_roll+17}'] = '-'

            
        if columns_CA_2:
            average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+13}' for col_letter in columns_CA_2])}),1)"
            mySheet2[f'C{total_roll+18}'] = average_formula
        else:
            mySheet2[f'C{total_roll+18}'] = '-'
        
        if columns_CA_3:
            average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+13}' for col_letter in columns_CA_3])}),1)"
            mySheet2[f'C{total_roll+19}'] = average_formula
        else:
            mySheet2[f'C{total_roll+19}'] = '-'
            
        if columns_CA_4:
            average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+13}' for col_letter in columns_CA_4])}),1)"
            mySheet2[f'C{total_roll+20}'] = average_formula
        else:
            mySheet2[f'C{total_roll+20}'] = '-'

        if columns_CA_5:
            average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+13}' for col_letter in columns_CA_5])}),1)"
            mySheet2[f'C{total_roll+21}'] = average_formula
        else:
            mySheet2[f'C{total_roll+21}'] = '-'

        if(cosCount == 6):
            if columns_CA_6:
                average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+13}' for col_letter in columns_CA_6])}),1)"
                mySheet2[f'C{total_roll+22}'] = average_formula
            else:
                mySheet2[f'C{total_roll+22}'] = '-'
        
        map_CA_co_arr=[f'={mySheet2.title}!C{total_roll+17}',f'={mySheet2.title}!C{total_roll+18}',f'={mySheet2.title}!C{total_roll+19}',f'={mySheet2.title}!C{total_roll+20}',f'={mySheet2.title}!C{total_roll+21}',f'={mySheet2.title}!C{total_roll+22}']
        
        return map_CA_co_arr
    
    def cal_PPT(mySheet3, al_value):
        
        highest_marks = 10

        # Variable to store group size
        groupSize = 1

        # For finding the size of the group
        for row in range(3, 100):
            if mySheet3[f'G{row}'].value is None:
                groupSize += 1
            else:
                break
        
        
        condition = highest_marks * float(al_value) // 100
        print(al_value, condition)



        # Set headers and make them bold
        headers = ['H1', 'I1', 'J1', 'K1']
        titles = ["Count", f'Count >= {al_value}%', f'% Count ( >= {al_value}%)', f"AL (Based on >= {al_value}% Count)"]

        for header, title in zip(headers, titles):
            mySheet3[header] = title
            mySheet3[header].font = Font(bold=True)

        # Set column widths
        for col in ['H', 'I', 'J', 'K']:
            mySheet3.column_dimensions[col].width = 20

        num_groups = total_roll // groupSize
        remaining_students = total_roll % groupSize

        if remaining_students > 0:
            num_groups += 1

        # Process each group
        row_counter = 2

        for group_number in range(num_groups):
            end_row = row_counter + groupSize - 1

            # Merge cells, apply alignment, and set borders for columns H, I, J, K
            for col in ['H', 'I', 'J', 'K']:
                mySheet3.merge_cells(f'{col}{row_counter}:{col}{end_row}')
                cell = mySheet3[f'{col}{row_counter}']
                cell.alignment = Alignment(horizontal='center', vertical='center')

                for row in range(row_counter, end_row + 1):
                    mySheet3[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    mySheet3[f'{col}{row}'].border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000') if row == row_counter else Side(style=None),
                        bottom=Side(style='thin', color='000000') if row == end_row else Side(style=None)
                    )

            row_counter += groupSize

        groupStart = 2
        endGroup = groupStart + groupSize - 1

        for i in range(num_groups):
            # Set the value in the top-left cell of the merged range
            mySheet3[f'H{groupStart}'] = f'=COUNT(F{groupStart}:F{endGroup})'
            mySheet3[f'I{groupStart}'] = f'=COUNTIF(F{groupStart}:F{endGroup},">={condition}")'
            mySheet3[f'J{groupStart}'] = f'=ROUND((I{groupStart}/H{groupStart})*100,1)'
            mySheet3[f'K{groupStart}'] = f'=IF(J{groupStart}<60,1,IF(AND(J{groupStart}>59,J{groupStart}<70),2,IF(AND(J{groupStart}>69,J{groupStart}<80),3,4)))'

            groupStart += groupSize
            endGroup += groupSize

        # Add 4 blank rows before the new table
        start_row_for_new_table = row_counter + 4

        # Set the headers for the new table
        mySheet3[f'H{start_row_for_new_table}'] = "COs"
        mySheet3[f'I{start_row_for_new_table}'] = "Average AL"
        mySheet3[f'H{start_row_for_new_table}'].font = Font(bold=True)
        mySheet3[f'I{start_row_for_new_table}'].font = Font(bold=True)

        # Set the width for the new columns
        mySheet3.column_dimensions['H'].width = 20
        mySheet3.column_dimensions['I'].width = 20

        # Calculate and set the values for COs and Average AL
        cos = [1, 2, 3, 4, 5, 6]
        if(cosCount == 5):
            cos = [1, 2, 3, 4, 5]
        al_row_start = start_row_for_new_table + 1
        avg_al_cells = []
        for co in cos:
            # Calculate the average AL for each CO manually using SUMIF and COUNTIF
            mySheet3[f'H{al_row_start}'] = co
            avg_al_cells.append(f'I{al_row_start}')
            # Center the values and apply borders
            for col in ['H', 'I']:
                mySheet3[f'{col}{al_row_start}'].alignment = Alignment(horizontal='center', vertical='center')
                mySheet3[f'{col}{al_row_start}'].border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )

            al_row_start += 1


        # Apply borders to the header of the new table
        for col in ['H', 'I']:
            mySheet3[f'{col}{start_row_for_new_table}'].alignment = Alignment(horizontal='center', vertical='center')
            mySheet3[f'{col}{start_row_for_new_table}'].border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

        groupStart = 2
        endGroup = groupStart + groupSize

        # Columns lists
        columns_with_1 = []
        columns_with_2 = []
        columns_with_3 = []
        columns_with_4 = []
        columns_with_5 = []
        columns_with_6 = [] 

        for group_number in range(num_groups):
            cell = mySheet3[f"G{groupStart}"]
            if cell.value != 0:
                values = [int(val.strip()) for val in str(cell.value).split(',') if val.strip().isdigit()]
                if 1 in values:
                    columns_with_1.append(groupStart)
                if 2 in values:
                    columns_with_2.append(groupStart)
                if 3 in values:
                    columns_with_3.append(groupStart)
                if 4 in values:
                    columns_with_4.append(groupStart)
                if 5 in values:
                    columns_with_5.append(groupStart)
                if 6 in values:
                    columns_with_6.append(groupStart)
            
            groupStart += groupSize

        # Helper function to generate formula
        def generate_average_formula(column_list):
            if column_list:
                # Join the list of column numbers into a string without square brackets
                return f"=ROUND(AVERAGE({','.join([f'K{col}' for col in column_list])}),1)"
            else:
                return '-'

        # Calculate the average using Excel formula
        mySheet3[avg_al_cells[0]] = generate_average_formula(columns_with_1)
        mySheet3[avg_al_cells[1]] = generate_average_formula(columns_with_2)
        mySheet3[avg_al_cells[2]] = generate_average_formula(columns_with_3)
        mySheet3[avg_al_cells[3]] = generate_average_formula(columns_with_4)
        mySheet3[avg_al_cells[4]] = generate_average_formula(columns_with_5)
        if(cosCount==6):
            mySheet3[avg_al_cells[5]] = generate_average_formula(columns_with_6)
        
        if(cosCount==5):
            map_CA_co_arr=[f'={mySheet3.title}!{avg_al_cells[0]}',f'={mySheet3.title}!{avg_al_cells[1]}',f'={mySheet3.title}!{avg_al_cells[2]}',f'={mySheet3.title}!{avg_al_cells[3]}',f'={mySheet3.title}!{avg_al_cells[4]}']
        else:
            map_CA_co_arr=[f'={mySheet3.title}!{avg_al_cells[0]}',f'={mySheet3.title}!{avg_al_cells[1]}',f'={mySheet3.title}!{avg_al_cells[2]}',f'={mySheet3.title}!{avg_al_cells[3]}',f'={mySheet3.title}!{avg_al_cells[4]}',f'={mySheet3.title}!{avg_al_cells[5]}']
        
        return map_CA_co_arr
        


                
    
    #<----------------------Quiz------------------->
    # sheet3=workbook['Quiz']
    
    # for col in ['C', 'D', 'E', 'F', 'G','H','I','J','K','L']:
        
    #     sheet3[f'{col}{total_roll+3}'] = f'=COUNT({col}3:{col}{total_roll+2})'
    #     sheet3[f'{col}{total_roll+4}'] = f'=ROUND(AVERAGE({col}3:{col}{total_roll+2}), 0)'
    #     sheet3[f'{col}{total_roll+5}'] = f'=COUNTIF({col}3:{col}{total_roll+2}, ">={
    # 
    #  / 100 * 2}")'
    #     sheet3[f'{col}{total_roll+6}'] = f'=ROUND({sheet3[f"{col}{total_roll+5}"].coordinate} / {sheet3[f"{col}{total_roll+3}"].coordinate} * 100, 1)'
    #     sheet3[f'{col}{total_roll+7}'] = f'=COUNTIF({col}3:{col}{total_roll+2}, ">="&{col}{total_roll+4})'
    #     sheet3[f'{col}{total_roll+8}'] = f'=ROUND({sheet3[f"{col}{total_roll+7}"].coordinate} / {sheet3[f"{col}{total_roll+3}"].coordinate} * 100, 1)'
    #     sheet3[f'{col}{total_roll+9}'] = f'=IF({sheet3[f"{col}{total_roll+6}"].coordinate}<60, 1, IF(AND({sheet3[f"{col}{total_roll+6}"].coordinate}>59, {sheet3[f"{col}{total_roll+6}"].coordinate}<70), 2, IF(AND({sheet3[f"{col}{total_roll+6}"].coordinate}>69, {sheet3[f"{col}{total_roll+6}"].coordinate}<80), 3, 4)))'
    
    # columns_QZ_1 = []
    # columns_QZ_2 = []
    # columns_QZ_3 = []
    # columns_QZ_4 = []
    # columns_QZ_5 = []
    # columns_QZ_6 = []
    
    # for col in ['C', 'D', 'E', 'F', 'G','H', 'I', 'J','K','L']:
    #     cell = sheet3[f"{col}2"]
    #     if cell.value :
    #         value3 = [int(val.strip()) for val in str(cell.value)[2:].split(',') if val.strip().isdigit()]   #This for with CO like CO1,2,3
    #         if 1 in value3:
    #             # Add the cell value to the list for calculation
    #             columns_QZ_1.append(col)
    #         if 2 in value3:
    #             # Add the cell value to the list for calculation
    #             columns_QZ_2.append(col)
    #         if 3 in value3:
    #             # Add the cell value to the list for calculation
    #             columns_QZ_3.append(col)
    #         if 4 in value3:
    #             # Add the cell value to the list for calculation
    #             columns_QZ_4.append(col)
    #         if 5 in value3:
    #             # Add the cell value to the list for calculation
    #             columns_QZ_5.append(col)
    #         if 6 in value3:
    #             columns_QZ_6.append(col)

    # # Calculate the average using Excel formula
    # if columns_QZ_1:
    #     average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_1])}),1)"
    #     sheet3[f'D{total_roll+13}'] = average_formula
    # else:
    #     sheet3[f'D{total_roll+13}'] = '-'

        
    # if columns_QZ_2:
    #     average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_2])}),1)"
    #     sheet3[f'D{total_roll+14}'] = average_formula
    # else:
    #     sheet3[f'D{total_roll+14}'] = '-'
    
    # if columns_QZ_3:
    #     average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_3])}),1)"
    #     sheet3[f'D{total_roll+15}'] = average_formula
    # else:
    #     sheet3[f'D{total_roll+15}'] = '-'
        
    # if columns_QZ_4:
    #     average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_4])}),1)"
    #     sheet3[f'D{total_roll+16}'] = average_formula
    # else:
    #     sheet3[f'D{total_roll+16}'] = '-'

    # if columns_QZ_5:
    #     average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_5])}),1)"
    #     sheet3[f'D{total_roll+17}'] = average_formula
    # else:
    #     sheet3[f'D{total_roll+17}'] = '-'

    # if columns_QZ_6:
    #     average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter}{total_roll+9}' for col_letter in columns_QZ_6])}),1)"
    #     sheet3[f'D{total_roll+18}'] = average_formula
    # else:
    #     sheet3[f'D{total_roll+18}'] = '-'
    
    
    # map_quiz_co_arr=[f'=Quiz!D{total_roll+13}',f'=Quiz!D{total_roll+14}',f'=Quiz!D{total_roll+15}',f'=Quiz!D{total_roll+16}',f'=Quiz!D{total_roll+17}',f'=Quiz!D{total_roll+18}']
    
    my_CA1_Co_arr=call_CA(workbook['CA1'], al_values_temp[0])
    my_CA2_Co_arr=call_CA(workbook['CA2'], al_values_temp[1])
    if 'CA3' in workbook.sheetnames:
        my_CA3_Co_arr=call_CA(workbook['CA3'], al_values_temp[2])
        
    #<----------------------Survey------------------->
    
    sheet4=workbook['Survey']
    
    for col in ['G','H','I','J','K','L']:
        
        sheet4[f'{col}{total_roll+4}'] = f'=COUNT({col}2:{col}{total_roll+1})'
        sheet4[f'{col}{total_roll+5}'] = f'=COUNTIF({col}2:{col}{total_roll+1}, ">=4")'
        sheet4[f'{col}{total_roll+6}'] = f'=ROUND(({col}{total_roll+5}/{col}{total_roll+4}*100), 1)'
        sheet4[f'{col}{total_roll+8}'] = f'=IF({col}{total_roll+6}<60,1,IF(AND({col}{total_roll+6}>59,{col}{total_roll+6}<70),2,IF(AND({col}{total_roll+6}>69,{col}{total_roll+6}<80),3,4)))'
        
    map_survey_co_arr=[f'={sheet4.title}!G{total_roll+8}',f'={sheet4.title}!H{total_roll+8}',f'={sheet4.title}!I{total_roll+8}',f'={sheet4.title}!J{total_roll+8}',f'={sheet4.title}!K{total_roll+8}',f'={sheet4.title}!L{total_roll+8}']
    
    #<----------------------Attainment------------------->
    sheet5=workbook['Attainment']
    
    # sheet5[f'B{21+i}']=f'IF({map_midsem_co_arr[i]}="-"," ","✓")'
    
    # Remember for nptel and Quiz we have to make different options also change in template
    attainmentEnd = 6
    if(cosCount==5):
        attainmentEnd=5
    for i in range(0,attainmentEnd):
        sheet5[f'B{33+i}']=map_midsem_co_arr[i]
        sheet5[f'B{21+i}'] = f'=IF({map_midsem_co_arr[i][1:] if map_midsem_co_arr[i].startswith("=") else map_midsem_co_arr[i]}="-"," ","✓")'

    
    for i in range(0, attainmentEnd):
        sheet5[f'C{33+i}']=my_CA1_Co_arr[i]
        sheet5[f'C{21+i}']=f'=IF({my_CA1_Co_arr[i][1:] if my_CA1_Co_arr[i].startswith("=") else my_CA1_Co_arr[i]}="-"," ","✓")'
        
    for i in range(0,attainmentEnd):
        sheet5[f'D{33+i}']=my_CA2_Co_arr[i]
        # value = my_CA2_Co_arr[i]
        # print(f"Setting D{33+i} to {sheet5[f'D{33+i}'].value}")
        
        sheet5[f'D{21+i}']=f'=IF({my_CA2_Co_arr[i][1:] if my_CA2_Co_arr[i].startswith("=") else my_CA2_Co_arr[i]}="-"," ","✓")'
    
    if 'CA3' in workbook.sheetnames:
        for i in range(0,attainmentEnd):
            sheet5[f'E{33+i}']=my_CA3_Co_arr[i]
            sheet5[f'E{21+i}']=f'=IF({my_CA3_Co_arr[i][1:] if my_CA3_Co_arr[i].startswith("=") else my_CA3_Co_arr[i]}="-"," ","✓")'
        
        for i in range(0,attainmentEnd):
            sheet5[f'F{33+i}']=map_endsem_co_arr[i]
            sheet5[f'F{21+i}']=f'=IF({map_endsem_co_arr[i][1:] if map_endsem_co_arr[i].startswith("=") else map_endsem_co_arr[i]}="-"," ","✓")'
            
        for i in range(0,attainmentEnd):
            sheet5[f'H{33+i}']=map_survey_co_arr[i]
            sheet5[f'G{21+i}']=f'=IF({map_survey_co_arr[i][1:] if map_survey_co_arr[i].startswith("=") else map_survey_co_arr[i]}="-"," ","✓")'
            
        for i in range(0,attainmentEnd):
            sheet5[f'E{43+i}']=map_survey_co_arr[i] 
            
        for i in range(0,attainmentEnd):
            # sheet5[f'G{33+i}']=f'=ROUND(0.7*F{33+i}+0.3*(AVERAGE(B{33+i},C{33+i},D{33+i},E{33+i})),1)'
            sheet5[f'G{33+i}']=f'=IF(AND(F{33+i}="-", COUNTIF(B{33+i}:E{33+i}, "-")=3), "-", IF(F{33+i}="-", ROUND(0.3*AVERAGE(B{33+i},C{33+i},D{33+i},E{33+i}), 1), IF(COUNTIF(B{33+i}:E{33+i}, "-")=4, ROUND(0.7*F{33+i}, 1), ROUND(0.7*F{33+i}+0.3*(AVERAGE(B{33+i},C{33+i},D{33+i},E{33+i})),1))))'
        for i in range(0,attainmentEnd):
            sheet5[f'D{43+i}']=sheet5[f'G{33+i}'].value        
       
    else:    
        for i in range(0,attainmentEnd):
            sheet5[f'E{33+i}']=map_endsem_co_arr[i]
            sheet5[f'E{21+i}']=f'=IF({map_endsem_co_arr[i][1:] if map_endsem_co_arr[i].startswith("=") else map_endsem_co_arr[i]}="-"," ","✓")'
                
        for i in range(0,attainmentEnd):
            sheet5[f'G{33+i}']=map_survey_co_arr[i]
            sheet5[f'F{21+i}']=f'=IF({map_survey_co_arr[i][1:] if map_survey_co_arr[i].startswith("=") else map_survey_co_arr[i]}="-"," ","✓")'
            
        for i in range(0,attainmentEnd):
            sheet5[f'E{43+i}']=map_survey_co_arr[i] 
            
            
        for i in range(0,attainmentEnd):
            # sheet5[f'F{33+i}']=f'=ROUND(0.7*E{33+i}+0.3*(AVERAGE(B{33+i},C{33+i},D{33+i})),1)'
            sheet5[f'F{33+i}']=f'=IF(AND(E{33+i}="-", COUNTIF(B{33+i}:D{33+i}, "-")=3), "-", IF(E{33+i}="-", ROUND(0.3*AVERAGE(B{33+i},C{33+i},D{33+i}), 1), IF(COUNTIF(B{33+i}:D{33+i}, "-")=3, ROUND(0.7*E{33+i}, 1), ROUND(0.7*E{33+i}+0.3*(AVERAGE(B{33+i},C{33+i},D{33+i})),1))))'
            
        for i in range(0,attainmentEnd):
            sheet5[f'D{43+i}']=sheet5[f'F{33+i}'].value        
        
    if(cosCount == 6):    
        map_co_arr=[f"={sheet5.title}!D43",f"={sheet5.title}!D44",f"={sheet5.title}!D45",f"={sheet5.title}!D46",f"={sheet5.title}!D47",f"={sheet5.title}!D48"]
    else:
        map_co_arr=[f"={sheet5.title}!D43",f"={sheet5.title}!D44",f"={sheet5.title}!D45",f"={sheet5.title}!D46",f"={sheet5.title}!D47"]

    # --  ---   -- PO Attainment --  ---  -- #
    sheet6=workbook['PO Attainment']

    start0 = 16
    start1 = 27
    start2 = 38
    column_array = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']

    for j in range(0, cosCount):
        for i in column_array:
            sheet6[f'{i}{start1+j}'] = map_co_arr[j]
            sheet6[f'{i}{start2+j}'] = f"=IF({i}{start0+j}=3, {i}{start1+j}, IF({i}{start0+j}=2, {i}{start1+j}*0.6, IF({i}{start0+j}=1, {i}{start1+j}*0.4, 0)))"
    print("type bataooooo")
    print(type(sheet6[f'{i}{start1+j}'].value))
    for j in range(0, cosCount):
        for i in column_array:
            print(f"2nd tabel {sheet6[f'{i}{start1+j}'].value}")
            print(sheet6[f'{i}{start0+j}'].value)
            if(sheet6[f'{i}{start0+j}'].value is None):
                 print("hehe")
                 sheet6[f'{i}{start2+j}'].value = ""

    for j in column_array:
        sheet6[f'{j}44'] = f"=ROUND(AVERAGE({j}38:{j}43),1)"

    ## sheet Done #
    # workbook.save('C:/Users/saira/Downloads/calTemplate.xlsx')
    selectedPath = filedialog.askdirectory()
    downloadCalculate = f'{selectedPath}/Calculated_{file_name_only}.xlsx'

    workbook.save(downloadCalculate)
    CTkMessagebox(message=f"Calculated excel sheet downloaded successfully at {downloadCalculate}.",icon="check", option_1="OK")


# EMAIL Part - need helps 

    # def send_email(sender_email, sender_password, recipient_email, subject, body, file_path):
    #     try:
    #         # Create a multipart message
    #         message = MIMEMultipart()
    #         message['From'] = sender_email
    #         message['To'] = recipient_email
    #         message['Subject'] = subject

    #         # Attach the email body
    #         message.attach(MIMEText(body, 'plain'))

    #         # Attach the file
    #         with open(file_path, "rb") as attachment:
    #             part = MIMEBase("application", "octet-stream")
    #             part.set_payload(attachment.read())

    #         encoders.encode_base64(part)
    #         part.add_header(
    #             "Content-Disposition",
    #             f"attachment; filename={os.path.basename(file_path)}"
    #         )
    #         message.attach(part)

    #         # Connect to the SMTP server and send the email
    #         with smtplib.SMTP('smtp.gmail.com', 587) as server:
    #             server.starttls()
    #             server.login(sender_email, sender_password)
    #             server.sendmail(sender_email, recipient_email, message.as_string())
    #         print("Email sent successfully!")

    #     except Exception as e:
    #         print(f"Error sending email: {e}")

    # # Main processing code
    # def process_and_send_file():

    #     # Notify the user
    #     print(f"Calculated excel sheet downloaded successfully at {downloadCalculate}.")

    #     # Input recipient email and other email details
    #     email_address = receiversEmail
    #     sender_email = "copoautomation@gmail.com"  # Replace with your email
    #     sender_password = "jbzs zfrc ibrg nelp"      # Replace with your email's app password
    #     subject = "Processed Excel File"
    #     body = f"Please find the attached processed Excel file - Calculated_{file_name_only}"

    #     # Send the file via email
    #     send_email(sender_email, sender_password, email_address, subject, body, downloadCalculate)

    # # Call the function
    # process_and_send_file()
   
    
    
    
    
    
# #cal_sheet()