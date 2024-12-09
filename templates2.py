import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook

# User inputs
subject = input("Enter the subject name: ")
total_roll = int(input("Enter the total number of roll numbers: "))
LabTarget = int(input("Enter the Lab Target: "))
groupSize = int(input("Enter the size of group: "))
criteria = int(input("Enter the number of criteria for marks: "))

critList = []
loList = []

for i in range(criteria):
    crit = input(f"Enter criteria {i + 1}: ")
    lo = input(f"Enter the LO for criteria {i + 1}: ")
    critList.append(crit)
    loList.append(lo)

# Create workbook and sheet
workbook = Workbook()
sheet1 = workbook.active
sheet1.title = "Lab"

# Styling for title
sheet1.merge_cells('A1:E1')
sheet1['A1'] = f"{subject} Lab Work"
sheet1['A1'].font = Font(size=14, bold=True)
sheet1['A1'].alignment = Alignment(horizontal='center')

# Add Target and headers
sheet1['A2'] = f"Target={LabTarget}"
sheet1['A3'] = "Group No."
sheet1['B3'] = "Roll No."
sheet1['C3'] = "Name of Student"
sheet1['D3'] = "Project Name"
sheet1['A4'] = "LOs Mapped"

# Adjust column width for Name column
sheet1.column_dimensions['C'].width = 30

# Group students based on group size
startCell = 5
groupCount = 1

# We will iterate over the total number of students, assign them to groups
for roll_no in range(1, total_roll + 1):
    current_row = startCell + roll_no - 1

    # Determine if we need to assign a new group number
    if (roll_no - 1) % groupSize == 0:
        sheet1[f"A{current_row}"] = groupCount  # Group number
        sheet1.merge_cells(f'A{current_row}:A{min(current_row + groupSize - 1, startCell + total_roll - 1)}')  # Merge group cells
        sheet1.merge_cells(f'D{current_row}:D{min(current_row + groupSize - 1, startCell + total_roll - 1)}')  # Merge project name cells
        groupCount += 1

    # Add roll number for each student
    sheet1[f"B{current_row}"] = roll_no
    sheet1[f"C{current_row}"] = f"Student {roll_no}"  # Placeholder for student names

# Adding criteria
for i in range(criteria):
    sheet1.cell(row=3, column=5 + i, value=critList[i])
    sheet1.cell(row=4, column=5 + i, value=loList[i])

# Apply borders to the entire table
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
sheet1[f'A{current_row+2}'] = f"Count>={LabTarget}"
sheet1[f'A{current_row+3}'] = f"%Count"
sheet1[f'A{current_row+4}'] = "AL"
for row in sheet1.iter_rows(min_row=3, max_row=startCell + total_roll - 1 + 4, min_col=1, max_col=5 + criteria):
    for cell in row:
        cell.border = thin_border

print(f'C{current_row}')
print(f'A{current_row+2}')
print(f'A{current_row+3}')
print(f'A{current_row+4}')


# Save the workbook
workbook.save(f"{subject}_Lab.xlsx")
