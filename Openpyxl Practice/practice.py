'''
from openpyxl import *
from openpyxl.utils import *
wb = load_workbook('Grades.xlsx')

ws = wb.active
print(ws)

print(ws['A1'].value)
print(ws['A2'].value)
print(ws['A3'].value)

ws['A3'].value = 'Eren'
ws['A4'] ='Mikasa'


#Accessing other sheets
print(wb.sheetnames)
ws = wb['Test Result'] # Just put sheet name

wb.create_sheet('Test') # Create new sheet with specified name

print(wb.sheetnames)
ws.title = 'Test Result'
ws.append(['Tim', 'is', 'Great','!'])


# Accessing cells 

for row in range(1,11):
    for col in range(1,5):
        char = get_column_letter(col) #we can add 65+col which will give letters
        print(ws[char+str(row)].value)

ws = wb['Test1']

# Merging cells
ws.merge_cells("A1:D1")
ws.unmerge_cells("A1:D1")

# Insert and Emptying cells
ws = wb['Test Result']
# ws.insert_rows(5) #inserting an empty row
ws.delete_rows(5)

# Copying and Moving cells

ws.move_range('A1:D11', rows=0, cols=0)
wb.save('Grades.xlsx')

'''

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data["Joe"].keys())
ws.append(headings)

for person in data:
    grades = [person] + list(data[person].values())
    ws.append(grades)

for col in range(2, len(data['Joe'])+2):
    char = get_column_letter(col)
    ws[char+ '7'] = f"=SUM({char+'2'}:{char+'6'})/{len(data)}"

wb.save("NewGrades.xlsx")