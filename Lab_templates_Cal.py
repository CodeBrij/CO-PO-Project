import re
from openpyxl.styles import *
from openpyxl import *

# Load the workbook and sheet
workbook = load_workbook('Lab_Records.xlsx')
sheet = workbook['Lab']

# Extract target value from cell A2
target = sheet['A2'].value
match = re.search(r'\d+', target)
if match:
    targetvalue = int(match.group())

# Extract expCount value from cell B2
expCount = sheet['B2'].value
match = re.search(r'\d+', expCount)
if match:
    expValue = int(match.group())

# Find the last non-empty row
i = 10
while sheet[f'A{i}'].value is not None:
    i += 1

# Last roll number found
countRoll = i - 1

# Starting column for processing
j = 3
for col in range(j, j + expValue):
    # Assign formula for COUNTIF
    sheet.cell(row=i, column=col, value=f"=COUNTIF(A4:A{countRoll + 4},\">={5 * targetvalue / 100}\")")
    # Assign formula for ROUND percentage
    sheet.cell(row=i + 1, column=col, value=f"=ROUND(({sheet.cell(row=i, column=col).coordinate}/{targetvalue})*100, 1)")
    # Assign formula for grading
    sheet.cell(row=i + 2, column=col, value=(
        f"=IF({sheet.cell(row=i + 1, column=col).coordinate}<60,1,"
        f"IF(AND({sheet.cell(row=i + 1, column=col).coordinate}>=60,"
        f"{sheet.cell(row=i + 1, column=col).coordinate}<70),2,"
        f"IF(AND({sheet.cell(row=i + 1, column=col).coordinate}>=70,"
        f"{sheet.cell(row=i + 1, column=col).coordinate}<80),3,4)))"
    ))

# Initialize LO counts
loCount = int(input("Number of LOs = "))
LOs = [0] * (loCount + 1)

# Process LOs based on grading
for col in range(3, expValue + 3):
    value = sheet.cell(row=5, column=col).value
    if value in range(1, loCount + 1):
        LOs[value] += value

# Output LO counts
print(LOs)

# Find the row to start adding the LOs-AL table (below the calculated data)
start_row = i + 5  # Leave some space after calculations

# Insert the headers for the table
sheet.cell(row=start_row, column=1, value="LOs")
sheet.cell(row=start_row, column=2, value="AL")

# Dictionary to store LO values and their corresponding columns
lo_columns = {}

# Map each unique LO in row 5 with the columns where they occur
for col in range(3, expValue + 3):
    lo_value = sheet.cell(row=5, column=col).value
    if lo_value not in lo_columns:
        lo_columns[lo_value] = []
    lo_columns[lo_value].append(col)

# Insert LOs and calculate AL (average) for each LO
for index, lo in enumerate(lo_columns.keys(), start=start_row + 1):
    # Insert LO value in the first column
    sheet.cell(row=index, column=1, value=lo)
    
    # Calculate the average of all the values for this LO across all mapped columns
    cols_range = ','.join([f"A{countRoll + 4}" for col in lo_columns[lo]])
    sheet.cell(row=index, column=2, value=f"=AVERAGE({cols_range})")

# Save the workbook after changes
workbook.save('Lab_Records_Updated.xlsx')
