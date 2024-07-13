import openpyxl
import os
from openpyxl.styles import *
# Get the path of the current directory
current_dir = os.path.dirname(__file__)

# Concatenate the file path with the "Project" folder and the file name
file_path = os.path.join(current_dir, "test.xlsx")

# Load the workbook using the updated file path
workbook = openpyxl.load_workbook(file_path)
sheet=workbook['Midsem']

coChar = 'A'
coRow = '1'
co = "COs"
target = 52.5
# For loop for finding start row number of roll nos.
for row in range(1,10):
    char = 'A'
    if(str(sheet[f'{char+str(row)}'].value)==co):
        coChar = char
        coRow = row
        break    

coRow =int(coRow) + 1  # Gives the start rowIndex for RollNo.
print(sheet['A78'].value)

# For Loop for finding the Number of students

# for row in range(coRow,200):
#     char = 'A'
#     if(str(sheet[f'{char+str(row)}'].value)=="None" or str(sheet[f'{char+str(row)}'].value)==""):
#         rollNum = row - coRow + 1
#         break

# print(rollNum)
# rollNumRow = (rollNum + coRow - 2) # 70 in this case 

# Entering the number of COs
COsNumber = int(input("ENter the number of COs: "))
COArray = [None] * COsNumber

midsemCOArray = [None] * COsNumber
eseCOArray = [None] * COsNumber
surveyCOArray = [None] * COsNumber
caCOArray = [None] * COsNumber
quizCOArray = [None] * COsNumber
rollNumRow = 72 + 1 # Not using just kabhi mistake se do baar click hua toh wrong aa jayega

sheet[f'A{str(rollNumRow+1)}'] = 'Count(Attempted)'
sheet[f'A{str(rollNumRow+2)}'] = 'Average  Marks'
sheet[f'A{str(rollNumRow+3)}'] = f'Count( >={target}%)'
sheet[f'A{str(rollNumRow+4)}'] = f'% Count( >={target}%)'
sheet[f'A{str(rollNumRow+5)}'] = 'Count( >=Average Marks)'
sheet[f'A{str(rollNumRow+6)}'] = f'AL(Based on >={target}% Count)'

for col in range (2,16):
    column_letter = openpyxl.utils.get_column_letter(col)
    sheet[f'{column_letter+str(rollNumRow+1)}']=f'=COUNT({column_letter+str(coRow)}:{column_letter+str(rollNumRow)})'

for row in range (coRow,rollNumRow+1):    
    sheet[f'H{row}']=f'=ROUND(SUM(B{row}:G{row}),0)'

for row in range (coRow,rollNumRow+1):    
    sheet[f'K{row}']=f'=ROUND(SUM(I{row}:J{row}),0)'
    
for row in range (coRow,rollNumRow+1):    
    sheet[f'N{row}']=f'=ROUND(SUM(L{row}:M{row}),0)'
        
for row in range (coRow,rollNumRow+1):    
    sheet[f'O{row}']=f'=ROUND(SUM(H{row},K{row},N{row}),0)'

for col in range (2,16):
    column_letter = openpyxl.utils.get_column_letter(col)
    sheet[f'{column_letter+str(rollNumRow+2)}']=f'=ROUND(AVERAGE({column_letter+str(coRow)}:{column_letter+str(rollNumRow)}),0)'
 
for col in range (2,8):
    column_letter = openpyxl.utils.get_column_letter(col)
    sheet[f'{column_letter+str(rollNumRow+3)}']=f'=COUNTIF({column_letter+str(coRow)}:{column_letter+str(rollNumRow)},">={target/100 * 2}")'

for col in range (8,16):
    column_letter = openpyxl.utils.get_column_letter(col)
    sheet[f'{column_letter+str(rollNumRow+3)}']=f'=COUNTIF({column_letter+str(coRow)}:{column_letter+str(rollNumRow)},">={target/100 * 5}")'

for col in range (2,16):
    column_letter = openpyxl.utils.get_column_letter(col)
    sheet[f'{column_letter+str(rollNumRow+4)}']=f'=ROUND(({column_letter+str(rollNumRow+3)}/{column_letter+str(rollNumRow+1)})*100,1)'


for col in range (2,16):
    column_letter = openpyxl.utils.get_column_letter(col)
    sheet[f'{column_letter+str(rollNumRow+5)}']=f'=COUNTIF({column_letter+str(coRow)}:{column_letter+str(rollNumRow)},">="&{column_letter+str(rollNumRow+2)})'


for col in range (2,16):
    column_letter = openpyxl.utils.get_column_letter(col)
    sheet[f'{column_letter+str(rollNumRow+6)}']=f'=IF({column_letter+str(rollNumRow+4)}<60,1,IF(AND({column_letter+str(rollNumRow+4)}>59,{column_letter+str(rollNumRow+4)}<70),2,IF(AND({column_letter+str(rollNumRow+4)}>69,{column_letter+str(rollNumRow+4)}<80),3,4)))'
    

coTableRow = rollNumRow + 6 + 3

columns_with_1 = []
columns_with_2 = []
columns_with_3 = []
columns_with_4 = []
columns_with_5 = []
columns_with_6 = [] 

    
column_range = ['B', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'L', 'M']

# Iterate through cells in the specified column range
for column_letter in column_range:
    # Get the cell in row 2 corresponding to the column letter
    cell = sheet[f"{column_letter}6"]
    # Check if the cell has a value
    if cell.value :
        # values = [int(val.strip()) for val in str(cell.value).split(',') if val.strip().isdigit()]  #This for without CO like 1,2,3
        values = [int(val.strip()) for val in str(cell.value)[2:].split(',') if val.strip().isdigit()]   #This for with CO like CO1,2,3
        # Check if '1' is present in the list of values then 2 and 3,4and 5
        # print(values)
        if 1 in values:
            # Add the cell value to the list for calculation
            columns_with_1.append(column_letter)
        if 2 in values:
            # Add the cell value to the list for calculation
            columns_with_2.append(column_letter)
        if 3 in values:
            # Add the cell value to the list for calculation
            columns_with_3.append(column_letter)
        if 4 in values:
            # Add the cell value to the list for calculation
            columns_with_4.append(column_letter)
        if 5 in values:
            # Add the cell value to the list for calculation
            columns_with_5.append(column_letter)
        if(COsNumber==6):
            if 6 in values:
                columns_with_6.append(column_letter)

# Calculate the average using Excel formula
if columns_with_1:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter+str(rollNumRow+6)}' for col_letter in columns_with_1])}),1)"
    sheet[f'G{coTableRow+1}'] = average_formula
    print(sheet[f'G{coTableRow+1}'].internal_value)
    print(sheet[f'G{coTableRow+1}'].internal_value)
    print(sheet[f'G{coTableRow+1}'].internal_value)
    midsemCOArray[0] = sheet[f'G{coTableRow+1}'].value
else:
    sheet[f'G{coTableRow+1}'] = '-'
    midsemCOArray[0] = 0

    
if columns_with_2:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter+str(rollNumRow+6)}' for col_letter in columns_with_2])}),1)"
    sheet[f'G{coTableRow+2}'] = average_formula
    midsemCOArray[1] = sheet[f'G{coTableRow+2}'].value
else:
    sheet[f'G{coTableRow+2}'] = '-'
    midsemCOArray[1] = 0
 
if columns_with_3:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter+str(rollNumRow+6)}' for col_letter in columns_with_3])}),1)"
    sheet[f'G{coTableRow+3}'] = average_formula
    midsemCOArray[2] = sheet[f'G{coTableRow+3}'].value
else:
    sheet[f'G{coTableRow+3}'] = '-'
    midsemCOArray[2] = 0
    
if columns_with_4:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter+str(rollNumRow+6)}' for col_letter in columns_with_4])}),1)"
    sheet[f'G{coTableRow+4}'] = average_formula
    midsemCOArray[3] = sheet[f'G{coTableRow+4}'].value
else:
    sheet[f'G{coTableRow+4}'] = '-'
    midsemCOArray[3] = 0

if columns_with_5:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter+str(rollNumRow+6)}' for col_letter in columns_with_5])}),1)"
    sheet[f'G{coTableRow+5}'] = average_formula
    midsemCOArray[4] = sheet[f'G{coTableRow+5}'].value
else:
    sheet[f'G{coTableRow+5}'] = '-'
    midsemCOArray[4] = 0

if(COsNumber == 6):
    if columns_with_6:
        average_formula = f"=ROUND(AVERAGE({','.join([f'{col_letter+str(rollNumRow+6)}' for col_letter in columns_with_6])}),1)"
        sheet[f'G{coTableRow+6}'] = average_formula
        midsemCOArray[5] = sheet[f'G{coTableRow+6}'].value
    else:
        sheet[f'G{coTableRow+6}'] = '-'
        midsemCOArray[5] = 0
    
print(sheet[f'G{coTableRow+5}'].internal_value)
    
sheet[f'F{coTableRow}'] = 'CO'
sheet[f'G{coTableRow}'] = 'AL'

sheet[f'F{coTableRow+1}'] = 1
sheet[f'F{coTableRow+2}'] = 2
sheet[f'F{coTableRow+3}'] = 3
sheet[f'F{coTableRow+4}'] = 4
sheet[f'F{coTableRow+5}'] = 5
if(COsNumber == 6):
     sheet[f'F{coTableRow+6}'] = 6


# workbook.create_sheet('ESE')
# ESE Sheet
# ESE Sheet
# ESE Sheet

sheet1 = workbook['ESE']

for i in range(1,20):
    char = 'A'
    if(str(sheet1[f'{char+str(i)}'].value)=='1'):
        startRow = i - 1
        break

for i in range(startRow,100):
    char = 'A'
    if(str(sheet1[f'{char+str(i)}'].value)=='None'):
        endRow = i
        break

studentNum = endRow - startRow + 1

startRow = startRow + 1
endRow = endRow + 1

sheet1[f'A{str(endRow+1)}'] = 'Count(Attempted)'
sheet1[f'A{str(endRow+2)}'] = 'Average  Marks'
sheet1[f'A{str(endRow+3)}'] = f'Count( >={target}%)'
sheet1[f'A{str(endRow+4)}'] = f'% Count( >={target}%)'
sheet1[f'A{str(endRow+5)}'] = 'Count( >=Average Marks)'
sheet1[f'A{str(endRow+6)}'] = f'AL(Based on >={target}% Count)'

count = 0
sum_marks = 0
total_marks = 60
column_letter = 'B'

# Define cell locations and formulas
sheet1[f'B{str(endRow+1)}'] = f'=COUNT({column_letter}{startRow}:{column_letter}{endRow})'
sheet1[f'B{str(endRow+2)}'] = f'=ROUND(AVERAGE({column_letter}{startRow}:{column_letter}{endRow}), 0)'
sheet1[f'B{str(endRow+3)}'] = f'=COUNTIF({column_letter}{startRow}:{column_letter}{endRow}, ">={target / 100 * total_marks}")'
sheet1[f'B{str(endRow+4)}'] = f'=ROUND({sheet1[f"B{str(endRow+3)}"].coordinate} / {sheet1[f"B{str(endRow+1)}"].coordinate} * 100, 1)'
sheet1[f'B{str(endRow+5)}'] = f'=COUNTIF({column_letter}{startRow}:{column_letter}{endRow}, ">="&{column_letter+str(endRow+2)})'
sheet1[f'B{str(endRow+6)}'] = f'=IF({sheet1[f"B{str(endRow+4)}"].coordinate}<60, 1, IF(AND({sheet1[f"B{str(endRow+4)}"].coordinate}>59, {sheet1[f"B{str(endRow+4)}"].coordinate}<70), 2, IF(AND({sheet1[f"B{str(endRow+4)}"].coordinate}>69, {sheet1[f"B{str(endRow+4)}"].coordinate}<80), 3, 4)))'
workbook.save("test.xlsx")
for i in range(0, COsNumber):
    print(i)
    print(sheet1[f'B{str(endRow+6)}'].value)
    eseCOArray[i] = sheet1[f'B{str(endRow+6)}'].value
    
# print(&{column_letter+str(endRow+2)})



# CO Course survey 
# CO Course survey 
# CO Course survey 
surveySheet = workbook['Course Exit Survey']

char_Sur = 'A'
surRow = 1

for row in range (2,100):
    if(surveySheet[f'{char_Sur+str(row)}'].value is None):
            surRow = row
            break

columnLetter = []

for col in range (5,50):
    colLetter = openpyxl.utils.get_column_letter(col)
    if(surveySheet[f'{colLetter+'1'}'].value is not None):
        columnLetter.append(colLetter)
    else:
        break

surveySheet[f'{'D'+str(surRow)}'] = 'Total'
surveySheet[f'{'D'+str(surRow+1)}'] = 'SA + A Count'
surveySheet[f'{'D'+str(surRow+2)}'] = 'SA + A Percentage'
surveySheet[f'{'D'+str(surRow+3)}'] = 'CO Mapped'
surveySheet[f'{'D'+str(surRow+4)}'] = 'AL'

for index in range (0,len(columnLetter)):
    surveySheet[f'{columnLetter[index]+str(surRow)}'] = f'=COUNT({columnLetter[index]+'2'}:{columnLetter[index]+str(surRow-1)})'
    surveySheet[f'{columnLetter[index]+str(surRow+1)}'] = f'=COUNTIF({columnLetter[index]+'2'}:{columnLetter[index]+str(surRow-1)}, ">=4")'
    surveySheet[f'{columnLetter[index]+str(surRow+2)}'] = f'=ROUND(({columnLetter[index]+str(surRow+1)}/{columnLetter[index]+str(surRow)}*100), 1)'
    surveySheet[f'{columnLetter[index]+str(surRow+3)}'] = 'CO'+f'{str(index+1)}'
    surveySheet[f'{columnLetter[index]+str(surRow+4)}'] = f'=IF({columnLetter[index]+str(surRow+2)}<60,1,IF(AND({columnLetter[index]+str(surRow+2)}>59,{columnLetter[index]+str(surRow+2)}<70),2,IF(AND({columnLetter[index]+str(surRow+2)}>69,{columnLetter[index]+str(surRow+2)}<80),3,4)))'

for index in range (0,len(columnLetter)):
    surveyCOArray[index] = surveySheet[f'{columnLetter[index]+str(surRow+4)}'].value


#Quiz
#Quiz
#Quiz

# Assuming 'workbook' is already defined and opened
sheet = workbook['Quiz']

char_quiz = 'A'
quizRow = 1
target = 52.5

# Find the first empty row in column A
for row in range(1, 100):
    if sheet[f'{char_quiz}{row}'].value is None:
        quizRow = row
        break

# Merge cells in columns A and B for four rows starting at quizRow
for i in range(0, 4):
    sheet.merge_cells(f'A{quizRow + i}:B{quizRow + i}')

# Set the values in the merged cells
sheet[f'A{quizRow}'] = 'Count(Attempted)'
sheet[f'A{quizRow + 1}'] = f'Count(>={target}%)'
sheet[f'A{quizRow + 2}'] = f'% Count(>={target}%)'
sheet[f'A{quizRow + 3}'] = f'AL(Based on >={target}% count)'

rollNoRow = 0
COrow = 0

# Find the rows for Roll No and CO
for row in range(1, 4):
    if sheet[f'{char_quiz}{row}'].value and 'CO' in str(sheet[f'{char_quiz}{row}'].value):
        COrow = row + 1
        rollNoRow = row + 1
        break

# Merge cells for "Attainment Level" and "CO"
merge1 = f'C{quizRow + 6}'
merge2 = f'D{quizRow + 6}'
sheet.merge_cells(f'{merge1}:{merge2}')
sheet[f'C{quizRow + 6}'] = 'Attainment Level'
sheet[f'C{quizRow + 7}'] = 'CO'

# Set CO labels
for i in range(0, 6):
    sheet[f'C{quizRow + 8 + i}'] = f'CO{i + 1}'

columnLetter = []

# Find columns with non-empty headers
for col in range(3, 60):
    colLetter = openpyxl.utils.get_column_letter(col)
    if sheet[f'{colLetter}1'].value is not None:
        columnLetter.append(colLetter)
    else:
        break

# Set formulas in the appropriate cells
for index in range(0, len(columnLetter)):
    sheet[f'{columnLetter[index]}{quizRow}'] = f'=COUNTA({columnLetter[index]}{rollNoRow}:{columnLetter[index]}{quizRow - 1})'
    sheet[f'{columnLetter[index]}{quizRow + 1}'] = f'=COUNTIF({columnLetter[index]}{rollNoRow}:{columnLetter[index]}{quizRow - 1}, ">={(target/100)*1}")'
    sheet[f'{columnLetter[index]}{quizRow + 2}'] = f'=ROUND(({columnLetter[index]}{quizRow + 1}/{columnLetter[index]}{quizRow}*100), 1)'
    sheet[f'{columnLetter[index]}{quizRow + 3}'] = f'=IF({columnLetter[index]}{quizRow + 2}<60,1,IF(AND({columnLetter[index]}{quizRow + 2}>59,{columnLetter[index]}{quizRow + 2}<70),2,IF(AND({columnLetter[index]}{quizRow + 2}>69,{columnLetter[index]}{quizRow + 2}<80),3,4)))'

CO1, CO2, CO3, CO4, CO5, CO6 = [], [], [], [], [], []
COrow=2
# Extract CO column references
for col in columnLetter:
    print(f'{col}{COrow}')
    cell = sheet[f'{col}{COrow}']
    if cell.value:
        cell_value = str(cell.value)
        print(f"Processing column {col}, value: {cell_value}")  # Debug statement
        if 'CO' in cell_value:
            co_index = cell_value.index('CO')
            co_numbers = ''
            for char in cell_value[co_index + 2:]:
                if char.isdigit() or char == ',':
                    co_numbers += char
                else:
                    break
            values = [int(val.strip()) for val in co_numbers.split(',') if val.strip().isdigit()]
        else:
            values = [int(val.strip()) for val in cell_value.split(',') if val.strip().isdigit()]

        print(f"Extracted CO values: {values}")  # Debug statement

        if 1 in values:
            CO1.append(col)
        if 2 in values:
            CO2.append(col)
        if 3 in values:
            CO3.append(col)
        if 4 in values:
            CO4.append(col)
        if 5 in values:
            CO5.append(col)
        if 6 in values:
            CO6.append(col)

# print(f"CO1 columns: {CO1}")  # Debug statement
# print(f"CO2 columns: {CO2}")  # Debug statement
# print(f"CO3 columns: {CO3}")  # Debug statement
# print(f"CO4 columns: {CO4}")  # Debug statement
# print(f"CO5 columns: {CO5}")  # Debug statement
# print(f"CO6 columns: {CO6}")  # Debug statement

# Calculate average formulas for each CO
if CO1:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO1])}),1)"
    sheet[f'D{quizRow + 8}'] = average_formula
    quizCOArray[0] = sheet[f'D{quizRow + 8}'].value
else:
    sheet[f'D{quizRow + 8}'] = '-'
    quizCOArray[0] = 0

if CO2:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO2])}),1)"
    sheet[f'D{quizRow + 9}'] = average_formula
    quizCOArray[1] = sheet[f'D{quizRow + 9}'].value
else:
    sheet[f'D{quizRow + 9}'] = '-'
    quizCOArray[1] = 0

if CO3:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO3])}),1)"
    sheet[f'D{quizRow + 10}'] = average_formula
    quizCOArray[2] = sheet[f'D{quizRow + 10}'].value
else:
    sheet[f'D{quizRow + 10}'] = '-'
    quizCOArray[2] = 0

if CO4:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO4])}),1)"
    sheet[f'D{quizRow + 11}'] = average_formula
    quizCOArray[3] = sheet[f'D{quizRow + 11}'].value
else:
    sheet[f'D{quizRow + 11}'] = '-'
    quizCOArray[3] = 0

if CO5:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO5])}),1)"
    sheet[f'D{quizRow + 12}'] = average_formula
    quizCOArray[4] = sheet[f'D{quizRow + 12}'].value
else:
    sheet[f'D{quizRow + 12}'] = '-'
    quizCOArray[4] = 0

if CO6:
    average_formula = f"=ROUND(AVERAGE({','.join([f'{column}{quizRow + 3}' for column in CO6])}),1)"
    sheet[f'D{quizRow + 13}'] = average_formula
    quizCOArray[5] = sheet[f'D{quizRow + 13}']
else:
    sheet[f'D{quizRow + 13}'] = '-'
    quizCOArray[5] = 0

#CO Attainment
#CO Attainment
#CO Attainment


sheets = workbook.sheetnames
print(sheets)
if(sheets.__contains__("CO Attainment")):
    print("Available")
else:
    workbook.create_sheet("CO Attainment")

attainmentSheet = workbook["CO Attainment"]

#Adding heading and required info

attainmentSheet.merge_cells("A1:I1")
attainmentSheet['A1'] = "Vivekanand  Education Society's Institute  of Technology"
attainmentSheet.merge_cells("A2:I2")
attainmentSheet['A2'] = "Department of Information Technology"
attainmentSheet.merge_cells("A3:I3")
academicYear = int(input("Enter the current academic Year (for e.g. 2023-2024 write 2024)"))
acadYear = str(academicYear-1)
attainmentSheet['A3'] = "Academic Year : " + str(acadYear) + "-" + str(academicYear)
attainmentSheet.merge_cells('A4:I4')
attainmentSheet.merge_cells('A5:I5')
attainmentSheet['A5'] = "Subject :                                 Class: 						"
attainmentSheet.merge_cells('A6:I6')
attainmentSheet['A6'] = "Subject Teacher:                          Sem : 							"
attainmentSheet.merge_cells('A8:I8')
attainmentSheet.merge_cells('A9:I9')
attainmentSheet.merge_cells('A10:I10')
attainmentSheet.merge_cells('A11:I11')
attainmentSheet.merge_cells('A12:I12')
attainmentSheet.merge_cells('A13:I13')
attainmentSheet['A8'] = "Course  Outcomes(COs): Upon successful completion of this course , students will be  able  to:"


#creating table
#arrays r ready insert the values in table and done
attainmentSheet.merge_cells('A17:E17')
attainmentSheet.merge_cells('B18:D18')
attainmentSheet.merge_cells('B19:C19')
attainmentSheet['A17'] = "CO -Rubrics Mapping"
attainmentSheet.merge_cells('A18:A19')
attainmentSheet['A18'] = "Assessment"
attainmentSheet['A20'] = "COs"
for i in range(0, COsNumber):
    attainmentSheet[f'A{21+i}'] = f'CO{i}'
attainmentSheet['B18'] = "Direct Assessment"
attainmentSheet['B19'] = "Internal Assessment"
attainmentSheet['D19'] = "External Assessment"
attainmentSheet['B20'] = "Mid Term Test"
attainmentSheet['C20'] = "Continuous Assessment"
attainmentSheet['D20'] = "ESE (TH)"
attainmentSheet.merge_cells('E18:E19')
attainmentSheet['E20'] = "Course Exit Survey"

for i in range(0, len(midsemCOArray)):
    if(midsemCOArray != 0):
        attainmentSheet[f'B{21+i}'] = midsemCOArray[i]

for i in range(0, len(caCOArray)):
    if(caCOArray[i] != 0):
        attainmentSheet[f'C{21+i}'] = caCOArray[i]

for i in range(0, len(eseCOArray)):
    if(eseCOArray[i] != 0):
        attainmentSheet[f'D{21+i}'] = "x"

for i in range(0, len(surveyCOArray)):
    if(surveyCOArray[i] != 0):
        attainmentSheet[f'E{21+i}'] = 'x'

attainmentSheet.merge_cells('A28:F28')
attainmentSheet.merge_cells('A30:A31')
attainmentSheet.merge_cells('F30:F31')
attainmentSheet.merge_cells('B30:E30')
attainmentSheet.merge_cells('B31:C31')

attainmentSheet['A30'] = "Assessment"
attainmentSheet['B30'] = "Direct Assessment"
attainmentSheet['B31'] = "Internal Assessment"
attainmentSheet['D31'] = "External Assessment"
attainmentSheet['E31'] = "Attainment level"
attainmentSheet['F30'] = "Indirect Assessment"
attainmentSheet['A32'] = "COs"
attainmentSheet['B32'] = "Mid Term Test"
attainmentSheet['C32'] = "Continuous Assessment"
attainmentSheet['D32'] = "ESE (TH)"
attainmentSheet['E32'] = "70% (External)+30% ( Internal)"
attainmentSheet['F32'] = "Course Exit survey"

for i in range(0, COsNumber):
    attainmentSheet[f'A{33+i}'] = f'CO{i}'

for i in range(0, len(midsemCOArray)):
    if(midsemCOArray != 0):
        attainmentSheet[f'B{33+i}'] = midsemCOArray[i]

for i in range(0,len(caCOArray)):
    if(caCOArray[i] != 0):
        attainmentSheet[f'C{33+i}'] = caCOArray[i]

for i in range(0, len(eseCOArray)):
    if(eseCOArray[i] != 0):
        attainmentSheet[f'D{33+i}'] = eseCOArray[i]

for i in range(0,len(surveyCOArray)):
    if(surveyCOArray[i] != 0):
        attainmentSheet[f'F{33+i}'] = surveyCOArray[i]

for i in range(0,COsNumber):
    attainmentSheet[f'E{33+i}'] = f'=ROUND(0.7*D{33+i}+0.3*(AVERAGE(B{33+i},C{33+i})),1)'

attainmentSheet.merge_cells('A41:C41')
attainmentSheet['A43'] = "Course Outcomes"
attainmentSheet['B43'] = "Direct AL"
attainmentSheet['C43'] = "Indirect AL"

for i in range(0, COsNumber):
    attainmentSheet[f'A{44+i}'] = f'CO{i}'

for i in range(0, COsNumber):
    attainmentSheet[f'B{44+i}'] = attainmentSheet[f'E{33+i}'].value
    attainmentSheet[f'C{44+i}'] = attainmentSheet[f'F{33+i}'].value

workbook.save("Test.xlsx")

