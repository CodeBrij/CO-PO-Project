import re
from openpyxl.styles import *
from openpyxl import *

# Load the workbook and sheet
workbook = load_workbook('Lab_Records.xlsx')
sheet = workbook['Lab']

# Extract target value from cell A2
target = sheet['A2'].value
match = re.search(r'\d+(\.\d+)?', target)
if match:
    targetvalue = float(match.group())

# Extract expCount value from cell B2
expCount = sheet['B2'].value
match = re.search(r'\d+', expCount)
if match:
    total_exp = int(match.group())

# Find the last non-empty row
i = 10
while sheet[f'A{i}'].value is not None:
    i += 1

# Last roll number row found
endRow = i - 1
startRow = 6
startCol = 'C'
offset = total_exp

# Calculate the new column letter
endCol = chr(ord(startCol) + offset - 1)
total_roll = endRow - 5

#alphabet array
alphabets = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

print(startRow)
print(endRow)
print(startCol)
print(endCol)
print(total_exp)
print(total_roll)
# Calculate the values


# Starting column for processing
j = 3
i = endRow + 2
for col in range(j, j + total_exp):
    # Assign formula for COUNTIF
    sheet.cell(row=i, column=col, value=f"=COUNTIF({alphabets[col-1]}4:{alphabets[col-1]}{endRow},\">={10 * targetvalue / 100}\")")
    print(f"=COUNTIF({alphabets[col-1]}4:{alphabets[col-1]}{endRow},\">={10 * targetvalue / 100}")
#     # Assign formula for ROUND percentage
    sheet.cell(row=i + 1, column=col, value=f"=ROUND(({sheet.cell(row=i, column=col).coordinate}/{total_roll})*100, 1)")

      # Assign formula for grading
    sheet.cell(row=i + 2, column=col, value=(
        f"=IF({sheet.cell(row=i + 1, column=col).coordinate}<60,1,"
        f"IF(AND({sheet.cell(row=i + 1, column=col).coordinate}>=60,"
        f"{sheet.cell(row=i + 1, column=col).coordinate}<70),2,"
        f"IF(AND({sheet.cell(row=i + 1, column=col).coordinate}>=70,"
        f"{sheet.cell(row=i + 1, column=col).coordinate}<80),3,4)))"
    ))

avgCol = chr(ord(startCol) + offset)
print(avgCol)
for i in range(total_roll):
    sheet[f'{avgCol}{i+6}'] = f'=ROUND(AVERAGE({startCol}{i+6}:{endCol}{i+6}),1)'

los_dict = {}
    
for col in range(ord(startCol), ord(endCol) + 1):
        current_col_letter = chr(col)
        cell_value = sheet[f'{current_col_letter}5'].value
        
        if cell_value:  # If the cell has a value
            values = cell_value.split(',')
            for value in values:
                value = value.strip()  # Remove any leading/trailing whitespace
                if value.isnumeric():  # Ensure it's a numeric value
                    lo_key = f'LO{value}'
                    if lo_key not in los_dict:
                        los_dict[lo_key] = []
                    los_dict[lo_key].append(current_col_letter)
    
print(los_dict)
# Define the border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define the font style for bold text
bold_font = Font(bold=True)
new_row = endRow+6
# Set the headers
sheet[f'B{new_row}'] = "LOs"
sheet[f'C{new_row}'] = "AL"

# Apply the border and bold font to the headers
sheet[f'B{new_row}'].border = thin_border
sheet[f'B{new_row}'].font = bold_font

sheet[f'C{new_row}'].border = thin_border
sheet[f'C{new_row}'].font = bold_font

i = 1

# Iterate over each key in los_dict to create and write the formulas
for key in los_dict.keys():
    columns = los_dict[key]  # Get the list of columns for the current LO
    column_ranges = ','.join([f'{col}{endRow+4}' for col in columns])  # Create the range for AVERAGE formula
    sheet[f'B{new_row+1}'] = key
    sheet[f'C{new_row+1}'] = f'=ROUND(AVERAGE({column_ranges}),1)'  # AVERAGE formula

    # Add border to the cells
    sheet[f'B{new_row+1}'].border = thin_border
    sheet[f'C{new_row+1}'].border = thin_border
    
    new_row += 1  # Increment new_row for the next set of entries



# # Save the workbook after changes
workbook.save('Lab_Records_Updated.xlsx')